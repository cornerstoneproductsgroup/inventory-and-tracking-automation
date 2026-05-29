import argparse
import csv
import html
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin

from playwright.sync_api import Browser, Page, TimeoutError, sync_playwright


def expand_env_vars(value: str) -> str:
    pattern = re.compile(r"\$\{([A-Za-z_][A-Za-z0-9_]*)\}")

    def replacer(match: re.Match[str]) -> str:
        name = match.group(1)
        return os.environ.get(name, "")

    return pattern.sub(replacer, value)


def deep_expand_env(data):
    if isinstance(data, dict):
        return {k: deep_expand_env(v) for k, v in data.items()}
    if isinstance(data, list):
        return [deep_expand_env(x) for x in data]
    if isinstance(data, str):
        return expand_env_vars(data)
    return data


def require_string(config: Dict, path: List[str]) -> str:
    current = config
    for key in path:
        if key not in current:
            raise ValueError(f"Missing config value: {'.'.join(path)}")
        current = current[key]
    if not isinstance(current, str) or not current.strip():
        raise ValueError(f"Config value must be a non-empty string: {'.'.join(path)}")
    return current


@dataclass
class CsvRecord:
    po_number: str
    tracking_number: str
    quantity: str
    shipment_type: str


def _fedex_po_lookup_keys_from_cell(raw: str, min_digit_run: int) -> List[str]:
    """
    FedEx / Worldship reference cells may contain SKU text plus the Lowe's PO in one field.
    Index every whitespace token and every long digit run so Rithum's PO can match a subset.
    """
    s = (raw or "").strip()
    if not s:
        return []
    seen: set[str] = set()
    out: List[str] = []

    def add(key: str) -> None:
        t = key.strip()
        if not t or t in seen:
            return
        seen.add(t)
        out.append(t)

    add(s)
    pat = rf"\d{{{min_digit_run},}}"
    for token in re.split(r"\s+", s):
        add(token)
        for m in re.finditer(pat, token):
            add(m.group(0))
    for m in re.finditer(pat, s):
        add(m.group(0))
    return out


def _canonical_po_from_fedex_cell(raw: str, min_digit_run: int) -> str:
    """Prefer the longest numeric PO-like token for logging / CsvRecord.po_number."""
    s = (raw or "").strip()
    if not s:
        return ""
    pat = rf"\d{{{min_digit_run},}}"
    runs = re.findall(pat, s)
    if runs:
        return max(runs, key=len)
    first = s.split()[0] if s.split() else s
    return first.strip()


def _digit_runs_at_least(text: str, min_len: int) -> List[str]:
    """Contiguous digit runs of at least min_len (avoids treating '407024614 EGLAI1' as 4070246141)."""
    ml = max(4, int(min_len))
    return re.findall(rf"\d{{{ml},}}", text or "")


def _po_lookup_candidate_strings(raw: str, min_digit_run: int) -> List[str]:
    """Ordered unique strings to try against csv_index keys (Rithum PO / reference line)."""
    s = (raw or "").strip()
    if not s:
        return []
    seen: set[str] = set()
    out: List[str] = []

    def add(t: str) -> None:
        u = (t or "").strip()
        if u and u not in seen:
            seen.add(u)
            out.append(u)

    add(s)
    parts = s.split()
    if parts:
        add(parts[0])
        add(parts[-1])
    for run in _digit_runs_at_least(s, min_digit_run):
        add(run)
    return out


class LowesTrackingAutomation:
    def __init__(self, config: Dict):
        self.config = config
        self.base_selectors = config["rithum"]["selectors"]
        self.stats = {
            "orders_seen": 0,
            "orders_matched": 0,
            "orders_submitted": 0,
            "orders_skipped_no_match": 0,
            "orders_failed": 0,
            "invoice_batches_submitted": 0,
        }
        self.csv_index: Dict[str, CsvRecord] = {}
        self._fedex_reference_min_digits: int = 5
        self._fedex_rows_loaded: int = 0

    def get_enabled_workflows(self, workflow_filter: str) -> List[Tuple[str, Dict]]:
        rithum = self.config["rithum"]
        workflows = rithum.get("workflows")
        selected: List[Tuple[str, Dict]] = []

        if isinstance(workflows, dict) and workflows:
            preferred_order = ["ship_to_store", "ship_to_customer", "invoice"]
            names = [x for x in preferred_order if x in workflows] + [
                x for x in workflows.keys() if x not in preferred_order
            ]
            for name in names:
                workflow_cfg = workflows[name]
                if not isinstance(workflow_cfg, dict):
                    continue
                if not bool(workflow_cfg.get("enabled", True)):
                    continue
                if workflow_filter != "all" and workflow_filter != name:
                    continue
                selected.append((name, workflow_cfg))

        # Backward-compatible fallback: single workflow driven by rithum.orders_url
        if not selected and workflow_filter in ("all", "default"):
            selected.append(
                (
                    "default",
                    {
                        "orders_url": rithum["orders_url"],
                        "selectors": {},
                    },
                )
            )
        return selected

    def get_effective_selectors(self, workflow_cfg: Dict) -> Dict:
        merged = dict(self.base_selectors)
        overrides = workflow_cfg.get("selectors", {})
        if isinstance(overrides, dict):
            merged.update(overrides)
        return merged

    def _normalize_rithum_url(self, raw_url: str) -> str:
        value = raw_url.strip()
        if value.startswith("http://") or value.startswith("https://"):
            return value
        # Supports config values like "gotoOrderRealmForm.do?..."
        return urljoin("https://dsm.commercehub.com/dsm/", value)

    def _extract_url_from_onclick(self, onclick: str) -> Optional[str]:
        if not onclick:
            return None
        text = html.unescape(onclick.strip())
        # Matches patterns like:
        # document.location='/dsm/gotoOrderRealmForm.do?orderid=...&action=web_shipgs1';
        # document.location="/dsm/gotoOrderRealmForm.do?...";
        match = re.search(r"document\.location\s*=\s*['\"]([^'\"]+)['\"]", text)
        if not match:
            return None
        return match.group(1).strip()

    def _column_letter_to_index(self, column_letter: str) -> int:
        text = column_letter.strip().upper()
        if not text or not re.fullmatch(r"[A-Z]+", text):
            raise ValueError(f"Invalid column letter: '{column_letter}'")

        index = 0
        for char in text:
            index = index * 26 + (ord(char) - ord("A") + 1)
        return index - 1

    def _resolve_tracking_file(self, csv_config: Dict) -> Path:
        explicit_path = (csv_config.get("path") or "").strip()
        if explicit_path:
            candidate = Path(explicit_path)
            if candidate.is_file():
                return candidate
            raise FileNotFoundError(f"FedEx file not found at configured path: {candidate}")

        directory = Path(require_string(self.config, ["fedex_csv", "directory"]))
        base_name = require_string(self.config, ["fedex_csv", "base_name"])
        extensions = csv_config.get("allowed_extensions", [".csv", ".xlsx"])
        if not isinstance(extensions, list) or not extensions:
            raise ValueError("fedex_csv.allowed_extensions must be a non-empty list")

        for ext in extensions:
            ext_text = str(ext).strip()
            if not ext_text:
                continue
            if not ext_text.startswith("."):
                ext_text = f".{ext_text}"
            candidate = directory / f"{base_name}{ext_text}"
            if candidate.is_file():
                return candidate

        searched = ", ".join([str(directory / f"{base_name}{x}") for x in extensions])
        raise FileNotFoundError(f"FedEx tracking file not found. Looked for: {searched}")

    def _upsert_record(
        self,
        po_value: str,
        tracking_value: str,
        quantity_value: str,
        shipment_type_value: str,
    ) -> None:
        tracking = tracking_value.strip()
        if not tracking:
            return
        min_run = max(1, int(self._fedex_reference_min_digits))
        keys = _fedex_po_lookup_keys_from_cell(po_value, min_run)
        if not keys:
            return
        canonical = _canonical_po_from_fedex_cell(po_value, min_run)
        record = CsvRecord(
            po_number=canonical,
            tracking_number=tracking,
            quantity=quantity_value.strip(),
            shipment_type=shipment_type_value,
        )
        # Keep first record by default. If duplicates exist, use the first
        # so behavior stays predictable.
        for key in keys:
            self.csv_index.setdefault(key, record)

        self._fedex_rows_loaded += 1

    def _load_from_csv_by_header(self, file_path: Path, csv_config: Dict, shipment_type_value: str) -> None:
        delimiter = csv_config.get("delimiter", ",")
        po_col = csv_config["po_column"]
        tracking_col = csv_config["tracking_column"]
        quantity_col = csv_config.get("quantity_column")
        quantity_default = str(csv_config.get("quantity_value", "1"))

        with file_path.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file, delimiter=delimiter)
            if reader.fieldnames is None:
                raise ValueError("FedEx CSV is missing header row")

            required = [po_col, tracking_col]
            if quantity_col:
                required.append(quantity_col)
            missing = [x for x in required if x not in reader.fieldnames]
            if missing:
                raise ValueError(
                    "FedEx CSV missing required columns: "
                    + ", ".join(missing)
                    + f". Found columns: {', '.join(reader.fieldnames)}"
                )

            for row in reader:
                quantity = (row.get(quantity_col) or "").strip() if quantity_col else quantity_default
                self._upsert_record(
                    po_value=(row.get(po_col) or ""),
                    tracking_value=(row.get(tracking_col) or ""),
                    quantity_value=quantity or quantity_default,
                    shipment_type_value=shipment_type_value,
                )

    def _load_from_csv_by_letter(self, file_path: Path, csv_config: Dict, shipment_type_value: str) -> None:
        delimiter = csv_config.get("delimiter", ",")
        po_index = self._column_letter_to_index(csv_config["po_column_letter"])
        tracking_index = self._column_letter_to_index(csv_config["tracking_column_letter"])
        quantity_letter = (csv_config.get("quantity_column_letter") or "").strip()
        quantity_index = self._column_letter_to_index(quantity_letter) if quantity_letter else None
        quantity_default = str(csv_config.get("quantity_value", "1"))
        start_row = int(csv_config.get("data_starts_on_row", 2))

        with file_path.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.reader(file, delimiter=delimiter)
            for row_num, row in enumerate(reader, start=1):
                if row_num < start_row:
                    continue

                po = row[po_index] if po_index < len(row) else ""
                tracking = row[tracking_index] if tracking_index < len(row) else ""
                quantity = (
                    row[quantity_index] if quantity_index is not None and quantity_index < len(row) else quantity_default
                )
                self._upsert_record(
                    po_value=str(po),
                    tracking_value=str(tracking),
                    quantity_value=str(quantity),
                    shipment_type_value=shipment_type_value,
                )

    def _load_from_xlsx_by_letter(self, file_path: Path, csv_config: Dict, shipment_type_value: str) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError as ex:
            raise RuntimeError(
                "openpyxl is required for Excel FedEx files. Run: pip install openpyxl"
            ) from ex

        po_index = self._column_letter_to_index(csv_config["po_column_letter"])
        tracking_index = self._column_letter_to_index(csv_config["tracking_column_letter"])
        quantity_letter = (csv_config.get("quantity_column_letter") or "").strip()
        quantity_index = self._column_letter_to_index(quantity_letter) if quantity_letter else None
        quantity_default = str(csv_config.get("quantity_value", "1"))
        start_row = int(csv_config.get("data_starts_on_row", 2))
        sheet_name = (csv_config.get("sheet_name") or "").strip()

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_num < start_row:
                    continue

                po = row[po_index] if po_index < len(row) and row[po_index] is not None else ""
                tracking = (
                    row[tracking_index]
                    if tracking_index < len(row) and row[tracking_index] is not None
                    else ""
                )
                quantity = (
                    row[quantity_index]
                    if quantity_index is not None and quantity_index < len(row) and row[quantity_index] is not None
                    else quantity_default
                )
                self._upsert_record(
                    po_value=str(po),
                    tracking_value=str(tracking),
                    quantity_value=str(quantity),
                    shipment_type_value=shipment_type_value,
                )
        finally:
            workbook.close()

    def load_csv_index(self) -> None:
        csv_config = self.config["fedex_csv"]
        self._fedex_rows_loaded = 0
        self._fedex_reference_min_digits = int(csv_config.get("reference_match_min_digits", 5))
        shipment_type_value = csv_config["shipment_type_value"]
        source_file = self._resolve_tracking_file(csv_config)
        source_ext = source_file.suffix.lower()

        if source_ext == ".csv":
            if csv_config.get("po_column_letter") and csv_config.get("tracking_column_letter"):
                self._load_from_csv_by_letter(source_file, csv_config, shipment_type_value)
            else:
                self._load_from_csv_by_header(source_file, csv_config, shipment_type_value)
        elif source_ext in (".xlsx", ".xlsm"):
            self._load_from_xlsx_by_letter(source_file, csv_config, shipment_type_value)
        else:
            raise ValueError(
                f"Unsupported FedEx source extension '{source_ext}'. "
                "Use .csv, .xlsx, or .xlsm."
            )

        key_count = len(self.csv_index)
        unique_records = len({id(rec) for rec in self.csv_index.values()})
        print(
            f"Loaded {self._fedex_rows_loaded} FedEx row(s) "
            f"({unique_records} shipment record(s), {key_count} PO lookup key(s)) from {source_file}"
        )

    def launch_browser(self, playwright) -> Browser:
        browser_settings = self.config["browser"]
        return playwright.chromium.launch(
            headless=bool(browser_settings.get("headless", False)),
            slow_mo=int(browser_settings.get("slow_mo_ms", 0)),
        )

    def login(self, page: Page) -> None:
        rithum = self.config["rithum"]
        username = require_string(self.config, ["rithum", "username"])
        password = require_string(self.config, ["rithum", "password"])
        selectors = self.base_selectors
        delay_config = rithum.get("login_delays_ms", {})
        delay_after_email_continue = int(delay_config.get("after_email_continue", 1200))
        delay_after_password_continue = int(delay_config.get("after_password_continue", 1200))
        delay_before_profile_selector = int(delay_config.get("before_profile_selector", 1500))

        page.goto(self._normalize_rithum_url(rithum["login_url"]), wait_until="domcontentloaded")
        page.locator(selectors["username_input"]).fill(username)
        username_continue_selector = (
            selectors.get("username_continue_button")
            or selectors.get("login_button")
            or "button[type='submit']"
        )
        page.locator(username_continue_selector).first.click()
        if delay_after_email_continue > 0:
            page.wait_for_timeout(delay_after_email_continue)

        page.locator(selectors["password_input"]).first.wait_for(timeout=30000)
        page.locator(selectors["password_input"]).fill(password)
        password_continue_selector = (
            selectors.get("password_continue_button")
            or selectors.get("login_button")
            or "button[type='submit']"
        )
        page.locator(password_continue_selector).first.click()
        if delay_after_password_continue > 0:
            page.wait_for_timeout(delay_after_password_continue)

        profile_selector = (selectors.get("profile_selector") or "").strip()
        if profile_selector:
            try:
                if delay_before_profile_selector > 0:
                    page.wait_for_timeout(delay_before_profile_selector)
                page.locator(profile_selector).first.wait_for(timeout=20000)
                page.locator(profile_selector).first.click()
            except Exception:
                # Continue if profile selector does not appear for this login.
                pass
        self._wait_for_commercehub_logged_in(page, selectors)
        print("Logged into Rithum.")

    def _wait_for_commercehub_logged_in(self, page: Page, selectors: Dict) -> None:
        """
        Confirm CommerceHub login without requiring a Lowe's-only dashboard link.
        When Lowe's has no open orders the PID=lowes link may not appear; Depot-only
        days should still proceed to Depot / Special Orders workflows.
        """
        _fast = os.environ.get("COMMERCEHUB_CHAIN_FAST") == "1"
        per_try_ms = 6000 if _fast else 10000
        candidates: list[str] = []
        primary = (selectors.get("logged_in_ready") or "").strip()
        if primary:
            candidates.append(primary)
        candidates.extend(
            [
                "a[href*='gotoOpenOrders.do?PID=lowes']",
                "a[href*='gotoOpenOrders.do?PID=thehomedepot']",
                "a[href*='gotoOpenOrders.do?PID=thdso']",
                "a[href*='gotoOpenOrders.do?PID=']",
                "a[href*='gotoOrderRealmForm.do']",
                "a[href*='gotoHome.do']",
            ]
        )
        seen: set[str] = set()
        last_err: Exception | None = None
        for sel in candidates:
            if not sel or sel in seen:
                continue
            seen.add(sel)
            try:
                page.locator(sel).first.wait_for(state="visible", timeout=per_try_ms)
                return
            except Exception as exc:
                last_err = exc
                continue
        url = (page.url or "").lower()
        if "commercehub.com" in url and not any(
            x in url for x in ("login", "signin", "okta", "auth0", "microsoftonline")
        ):
            return
        if last_err is not None:
            raise last_err
        raise TimeoutError("Could not confirm CommerceHub login.")

    def get_order_links(self, page: Page, workflow_name: str, workflow_cfg: Dict, selectors: Dict) -> List[str]:
        orders_url = workflow_cfg.get("orders_url") or self.config["rithum"].get("orders_url")
        if not orders_url:
            raise ValueError(f"Missing orders_url for workflow '{workflow_name}'")
        max_orders = int(self.config["automation"].get("max_orders", 50))
        page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
        _fast = os.environ.get("COMMERCEHUB_CHAIN_FAST") == "1"
        link_timeout_ms = 6000 if _fast else 15000
        try:
            page.locator(selectors["order_links"]).first.wait_for(timeout=link_timeout_ms)
        except Exception:
            print(f"[{workflow_name}] No order links on orders page; skipping workflow.")
            return []

        handles = page.locator(selectors["order_links"])
        total = handles.count()
        links: List[str] = []
        seen: set[str] = set()
        for i in range(min(total, max_orders)):
            row = handles.nth(i)
            href = row.get_attribute("href")
            raw_url = (href or "").strip()
            if not raw_url:
                onclick = row.get_attribute("onclick") or ""
                raw_url = self._extract_url_from_onclick(onclick) or ""
            if not raw_url:
                continue
            full_url = urljoin(page.url, raw_url)
            if full_url in seen:
                continue
            seen.add(full_url)
            links.append(full_url)
        print(f"[{workflow_name}] Found {len(links)} order link(s).")
        return links

    def lookup_record(self, po_number: str) -> Optional[CsvRecord]:
        min_run = max(5, int(getattr(self, "_fedex_reference_min_digits", 5)))
        normalized_po = (po_number or "").replace("\u00a0", " ").strip()

        for cand in _po_lookup_candidate_strings(normalized_po, min_run):
            if cand in self.csv_index:
                return self.csv_index[cand]

        # Prefix match: Rithum reference starts with a FedEx key, then trailing words.
        normalized_upper = normalized_po.upper()
        for key, value in self.csv_index.items():
            key_text = key.strip()
            if not key_text:
                continue
            key_upper = key_text.upper()
            if not normalized_upper.startswith(key_upper):
                continue
            if len(normalized_upper) == len(key_upper):
                return value
            next_char = normalized_upper[len(key_upper)]
            if not next_char.isalnum():
                return value

        # FedEx cell may be "407024614 EGLAI1": match digit *runs*, not all digits concatenated.
        po_runs = _digit_runs_at_least(normalized_po, min_run)
        if not po_runs:
            collapsed = re.sub(r"\D", "", normalized_po)
            if collapsed:
                po_runs = [collapsed]
        if not po_runs:
            return None

        for key, value in self.csv_index.items():
            key_runs = _digit_runs_at_least(key, min_run)
            if not key_runs:
                continue
            for pr in po_runs:
                if pr in key_runs:
                    return value
                if pr.isdigit():
                    for kr in key_runs:
                        if not kr.isdigit():
                            continue
                        if int(kr) == int(pr):
                            return value
        return None

    def set_input_or_select(self, page: Page, selector: str, value: str) -> None:
        node_name = page.locator(selector).evaluate("el => el.tagName.toLowerCase()")
        if node_name == "select":
            try:
                page.locator(selector).select_option(value=value)
            except Exception:
                page.locator(selector).select_option(label=value)
        else:
            page.locator(selector).fill(value)

    def _extract_numeric_quantity(self, text: str) -> Optional[str]:
        match = re.search(r"(\d+(?:\.\d+)?)", text or "")
        if not match:
            return None
        return match.group(1)

    def _parse_remaining_qty_display(self, text: str) -> Optional[str]:
        """
        Ship-to-store (React) shows remaining like \"2 EA\" in div.css-10ndwee.
        Prefer the integer before EA, then fall back to any number in the string.
        """
        t = (text or "").strip()
        if not t:
            return None
        m = re.search(r"(\d+(?:\.\d+)?)\s*EA", t, re.IGNORECASE)
        if m:
            return m.group(1)
        return self._extract_numeric_quantity(t)

    def _remaining_qty_in_row_with_ship_input(self, qty_input_locator) -> Optional[str]:
        """
        Quantity Remaining (\"2 EA\") must come from the *same table row* as the ship-qty
        input. A global ``div.css-10ndwee`` matches many unrelated nodes, so pairing by
        index often reads the wrong div and yields 1 from the FedEx fallback.
        """
        try:
            axis_variants = (
                "ancestor::tr[1]",
                "ancestor::*[local-name()='tr'][1]",
                "ancestor::*[@role='row'][1]",
            )
            for ax in axis_variants:
                row = qty_input_locator.locator(f"xpath={ax}")
                if row.count() == 0:
                    continue
                r0 = row.first
                try:
                    mcell = r0.get_by_text(re.compile(r"\d+(?:\.\d+)?\s*EA", re.I))
                    if mcell.count() > 0:
                        t = mcell.first.inner_text().strip()
                        p = self._parse_remaining_qty_display(t)
                        if p:
                            return p
                except Exception:
                    pass
                for sub in ("[class*='10ndwee']", "div[class*='10ndwee']", "*[contains(@class,'10ndwee')]"):
                    cells = r0.locator(sub)
                    for j in range(min(cells.count(), 30)):
                        t = cells.nth(j).inner_text().strip()
                        if not t or "EA" not in t.upper():
                            continue
                        p = self._parse_remaining_qty_display(t)
                        if p:
                            return p
            return None
        except Exception:
            return None

    def _resolve_ship_quantity_for_line(
        self,
        qty_input_locator,
        rem_locator,
        line_idx: int,
        workflow_cfg: Dict,
        fallback: str,
    ) -> str:
        """Resolve ship qty for one line (ship-to-store may have many qty + \"N EA\" pairs)."""
        if not bool(workflow_cfg.get("use_quantity_remaining", False)):
            return fallback

        parsed = self._remaining_qty_in_row_with_ship_input(qty_input_locator)
        if parsed:
            return parsed

        if rem_locator is not None:
            try:
                n_rem = rem_locator.count()
                if n_rem > 0:
                    rem_idx = min(line_idx, n_rem - 1)
                    remaining_text = rem_locator.nth(rem_idx).inner_text().strip()
                    parsed = self._parse_remaining_qty_display(remaining_text)
                    if parsed:
                        return parsed
            except Exception:
                pass

        try:
            max_attr = qty_input_locator.get_attribute("max") or ""
            parsed = self._extract_numeric_quantity(max_attr)
            if parsed:
                return parsed
        except Exception:
            pass

        return fallback

    def _extract_order_id_from_href(self, href: str) -> Optional[str]:
        match = re.search(r"[?&]Hub_PO=(\d+)", href or "")
        if match:
            return match.group(1)
        return None

    def _quantity_from_remaining_cell(self, page: Page, remaining_id: str) -> Optional[str]:
        try:
            loc = page.locator(f"[id='{remaining_id}']")
            if loc.count() == 0:
                return None
            remaining_text = loc.first.inner_text().strip()
            return self._parse_remaining_qty_display(remaining_text)
        except Exception:
            return None

    def _resolve_line_quantity_from_input(
        self,
        page: Page,
        qty_input,
        fallback_quantity: str,
        *,
        order_id: Optional[str] = None,
    ) -> str:
        """
        Prefer CommerceHub per-line \"remaining\" qty over FedEx default (often 1).
        Shipped inputs usually have name=order(...).item(...).shipped but no id.
        """
        shipped_suffix = ".shipped"
        input_id = (qty_input.get_attribute("id") or "").strip()
        if input_id.startswith("order(") and input_id.endswith(shipped_suffix):
            rid = f"cell.line.{input_id[: -len(shipped_suffix)]}.remaining"
            parsed = self._quantity_from_remaining_cell(page, rid)
            if parsed:
                return parsed

        name = (qty_input.get_attribute("name") or "").strip()
        if name.startswith("order(") and name.endswith(shipped_suffix):
            rid = f"cell.line.{name[: -len(shipped_suffix)]}.remaining"
            parsed = self._quantity_from_remaining_cell(page, rid)
            if parsed:
                return parsed

        # Depot-style: td id contains order id, item index, and .remaining
        if order_id and name:
            m = re.search(r"\.item\(([^)]+)\)\.shipped$", name)
            if m:
                item_token = m.group(1)
                try:
                    rem = page.locator(
                        "xpath=//td[contains(@id,'order("
                        + order_id
                        + ")') and contains(@id,'.item("
                        + item_token
                        + ")') and contains(@id,'.remaining')]"
                    )
                    if rem.count() > 0:
                        parsed = self._parse_remaining_qty_display(rem.first.inner_text().strip())
                        if parsed:
                            return parsed
                except Exception:
                    pass

        max_attr = (qty_input.get_attribute("max") or "").strip()
        parsed_max = self._extract_numeric_quantity(max_attr)
        if parsed_max:
            return parsed_max
        return fallback_quantity

    def process_orders_in_list(
        self,
        page: Page,
        workflow_name: str,
        workflow_cfg: Dict,
        selectors: Dict,
        do_submit: bool,
    ) -> None:
        orders_url = workflow_cfg.get("orders_url") or self.config["rithum"].get("orders_url")
        if not orders_url:
            raise ValueError(f"Missing orders_url for workflow '{workflow_name}'")

        po_links_selector = selectors.get("po_links") or selectors.get("po_on_order_page")
        if not po_links_selector:
            raise ValueError(f"[{workflow_name}] Missing selector 'po_links' for list workflow")

        page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
        _fast = os.environ.get("COMMERCEHUB_CHAIN_FAST") == "1"
        if _fast:
            page.wait_for_timeout(200)
        run_until_stopped = bool(workflow_cfg.get("run_until_stopped", False))
        run_until_queue_empty = bool(workflow_cfg.get("run_until_queue_empty", False))
        idle_sleep_seconds = int(workflow_cfg.get("idle_sleep_seconds", 10))
        if _fast:
            idle_sleep_seconds = min(idle_sleep_seconds, 1)
        po_row_timeout_ms = 4500 if _fast else 15000

        while True:
            try:
                page.locator(po_links_selector).first.wait_for(timeout=po_row_timeout_ms)
            except Exception:
                print(f"[{workflow_name}] No order rows currently visible.")
                if run_until_queue_empty:
                    return
                if do_submit and run_until_stopped:
                    page.wait_for_timeout(idle_sleep_seconds * 1000)
                    page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
                    continue
                return

            max_orders = int(self.config["automation"].get("max_orders", 50))
            po_links = page.locator(po_links_selector)
            total = min(po_links.count(), max_orders)
            processed_for_submit = 0

            for i in range(total):
                po_link = po_links.nth(i)
                po_number = po_link.inner_text().strip()
                if not po_number:
                    continue

                href = po_link.get_attribute("href") or ""

                self.stats["orders_seen"] += 1
                print(f"[{workflow_name}] Detected PO: {po_number}")

                record = self.lookup_record(po_number)
                if record is None:
                    hub_po = self._extract_order_id_from_href(href)
                    if hub_po:
                        record = self.lookup_record(hub_po)
                if record is None:
                    self.stats["orders_skipped_no_match"] += 1
                    if self.config["automation"].get("skip_orders_without_csv_match", True):
                        print(f"[{workflow_name}] No CSV match found for PO '{po_number}'. Skipping.")
                        continue
                    raise ValueError(f"No CSV match found for PO '{po_number}'")

                order_id = self._extract_order_id_from_href(href)
                if not order_id:
                    self.stats["orders_failed"] += 1
                    print(f"[{workflow_name}] Could not parse order id for PO '{po_number}'.")
                    continue

                tracking_inputs = page.locator(
                    f"input[name^='order({order_id}).box('][name$='.trackingnumber']"
                )
                shipping_selects = page.locator(
                    f"select[name^='order({order_id}).box('][name$='.shippingmethod']"
                )
                qty_inputs = page.locator(
                    f"input[name^='order({order_id}).box('][name$='.shipped']"
                )

                if tracking_inputs.count() == 0 or shipping_selects.count() == 0 or qty_inputs.count() == 0:
                    self.stats["orders_failed"] += 1
                    print(
                        f"[{workflow_name}] Required inputs not found for PO '{po_number}' "
                        f"(order id {order_id})."
                    )
                    continue

                for idx in range(tracking_inputs.count()):
                    tracking_inputs.nth(idx).fill(record.tracking_number)

                shipment_value = str(workflow_cfg.get("shipment_type_value_override") or record.shipment_type)
                for idx in range(shipping_selects.count()):
                    select = shipping_selects.nth(idx)
                    try:
                        select.select_option(value=shipment_value)
                    except Exception:
                        select.select_option(label="FedEx Ground")

                for idx in range(qty_inputs.count()):
                    qty_input = qty_inputs.nth(idx)
                    line_quantity = self._resolve_line_quantity_from_input(
                        page, qty_input, record.quantity, order_id=order_id
                    )
                    qty_input.fill(line_quantity)

                self.stats["orders_matched"] += 1
                processed_for_submit += 1
                print(
                    f"[{workflow_name}] Filled PO {po_number} "
                    f"(tracking={record.tracking_number}, ship_method={shipment_value})."
                )

            if do_submit and processed_for_submit > 0:
                page.locator(selectors["submit_button"]).click()
                success_selector = selectors.get("submit_success_banner")
                if success_selector:
                    banner_ms = 3500 if _fast else 10000
                    try:
                        page.locator(success_selector).first.wait_for(timeout=banner_ms)
                    except Exception:
                        page.wait_for_load_state("domcontentloaded")
                else:
                    page.wait_for_load_state("domcontentloaded")
                self.stats["orders_submitted"] += processed_for_submit
                print(f"[{workflow_name}] Submitted page with {processed_for_submit} order(s).")
                if run_until_stopped or run_until_queue_empty:
                    page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
                    continue
                return

            if do_submit and run_until_stopped:
                print(
                    f"[{workflow_name}] Nothing submitted this pass; "
                    f"retrying in {idle_sleep_seconds}s..."
                )
                page.wait_for_timeout(idle_sleep_seconds * 1000)
                page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
                continue

            if do_submit and run_until_queue_empty:
                print(f"[{workflow_name}] No orders filled this pass; exiting queue-empty mode.")
                return

            print(f"[{workflow_name}] Dry run mode: list values filled, submit skipped.")
            return

    def process_invoice_list(
        self,
        page: Page,
        workflow_name: str,
        workflow_cfg: Dict,
        selectors: Dict,
        do_submit: bool,
    ) -> None:
        orders_url = workflow_cfg.get("orders_url") or self.config["rithum"].get("orders_url")
        if not orders_url:
            raise ValueError(f"Missing orders_url for workflow '{workflow_name}'")

        select_all_sel = (selectors.get("invoice_select_all_button") or "").strip() or "input#checkall[name='checkall']"
        autofill_sel = (
            (selectors.get("invoice_autofill_buttons") or "").strip()
            or "input[type='button'][name$='.invoicenumber.autofill'][value='Auto Fill']"
        )
        submit_sel = (selectors.get("submit_button") or "").strip() or "input#confirmbtn[name='confirmbtn']"
        idle_after_submit_ms = int(workflow_cfg.get("idle_after_submit_ms", 800))
        delay_autofill_ms = int(workflow_cfg.get("delay_between_invoice_autofill_ms", 150))
        _fast = os.environ.get("COMMERCEHUB_CHAIN_FAST") == "1"
        if _fast:
            idle_after_submit_ms = min(idle_after_submit_ms, 250)
            delay_autofill_ms = min(delay_autofill_ms, 50)

        page.goto(self._normalize_rithum_url(orders_url), wait_until="domcontentloaded")
        if _fast:
            page.wait_for_timeout(200)

        invoice_select_timeout_ms = 6000 if _fast else 25000

        while True:
            try:
                page.locator(select_all_sel).first.wait_for(timeout=invoice_select_timeout_ms)
            except Exception:
                print(f"[{workflow_name}] Invoice UI not found (select all). Done or wrong page.")
                return

            autofill_count = page.locator(autofill_sel).count()
            if autofill_count == 0:
                print(f"[{workflow_name}] No invoice Auto Fill buttons on page; invoicing queue empty.")
                return

            print(
                f"[{workflow_name}] Invoice batch: Select All, then Auto Fill x{autofill_count}, "
                f"then Submit."
            )
            page.locator(select_all_sel).first.click()
            page.wait_for_timeout(150 if _fast else 300)

            for idx in range(autofill_count):
                page.locator(autofill_sel).nth(idx).click()
                if delay_autofill_ms > 0:
                    page.wait_for_timeout(delay_autofill_ms)

            if not do_submit:
                print(f"[{workflow_name}] Dry run: invoice actions skipped (no Submit).")
                return

            page.locator(submit_sel).first.click()
            page.wait_for_load_state("domcontentloaded")
            if idle_after_submit_ms > 0:
                page.wait_for_timeout(idle_after_submit_ms)
            self.stats["invoice_batches_submitted"] += 1
            print(f"[{workflow_name}] Submitted invoice batch #{self.stats['invoice_batches_submitted']}.")

    def process_order(
        self,
        page: Page,
        order_url: str,
        selectors: Dict,
        workflow_name: str,
        workflow_cfg: Dict,
        do_submit: bool,
    ) -> None:
        self.stats["orders_seen"] += 1
        print(f"\n[{workflow_name}] Processing order page: {order_url}")
        page.goto(order_url, wait_until="domcontentloaded")

        po_number = page.locator(selectors["po_on_order_page"]).inner_text().strip()
        if not po_number:
            raise ValueError("PO number was empty on the order page")

        print(f"[{workflow_name}] Detected PO: {po_number}")
        record = self.lookup_record(po_number)
        if record is None:
            hub_po = self._extract_order_id_from_href(order_url or "")
            if hub_po:
                record = self.lookup_record(hub_po)
        if record is None:
            self.stats["orders_skipped_no_match"] += 1
            if self.config["automation"].get("skip_orders_without_csv_match", True):
                print(f"[{workflow_name}] No CSV match found for PO '{po_number}'. Skipping.")
                return
            raise ValueError(f"No CSV match found for PO '{po_number}'")

        self.stats["orders_matched"] += 1
        print(
            f"CSV match found. Tracking={record.tracking_number}, "
            f"ShipmentType={record.shipment_type}, Quantity={record.quantity}"
        )

        self.set_input_or_select(page, selectors["tracking_input"], record.tracking_number)
        shipment_value = str(workflow_cfg.get("shipment_type_value_override") or record.shipment_type)
        self.set_input_or_select(page, selectors["shipment_type_input"], shipment_value)

        rem_sel = (selectors.get("quantity_remaining") or "").strip()
        rem_loc = page.locator(rem_sel) if rem_sel else None
        qty_loc = page.locator(selectors["quantity_input"])
        if qty_loc.count() == 0:
            raise ValueError("quantity_input selector matched no elements on the order page")
        for line_idx in range(qty_loc.count()):
            ship_quantity = self._resolve_ship_quantity_for_line(
                qty_loc.nth(line_idx),
                rem_loc,
                line_idx,
                workflow_cfg,
                record.quantity,
            )
            qty_loc.nth(line_idx).fill(str(ship_quantity))

        if self.config["automation"].get("pause_for_manual_review_before_submit", False):
            input("Press Enter to submit this order...")

        if do_submit:
            page.locator(selectors["submit_button"]).click()
            success_selector = selectors.get("submit_success_banner")
            if success_selector:
                _fast = os.environ.get("COMMERCEHUB_CHAIN_FAST") == "1"
                banner_ms = 3500 if _fast else 10000
                try:
                    page.locator(success_selector).first.wait_for(timeout=banner_ms)
                except Exception:
                    page.wait_for_load_state("domcontentloaded")
            self.stats["orders_submitted"] += 1
            print(f"[{workflow_name}] Submitted successfully.")
        else:
            print(f"[{workflow_name}] Dry run mode: values filled, submit skipped.")

    def run_workflows_after_login(self, page: Page, do_submit: bool, workflow_filter: str) -> None:
        """Run enabled workflows assuming CommerceHub login is already done on ``page``."""
        workflows = self.get_enabled_workflows(workflow_filter)
        if not workflows:
            raise ValueError(
                f"No enabled workflows matched filter '{workflow_filter}'. "
                "Check rithum.workflows config."
            )

        for workflow_name, workflow_cfg in workflows:
            wf_cfg = dict(workflow_cfg)
            if workflow_name == "ship_to_customer" and workflow_filter in (
                "all",
                "ship_to_customer",
            ):
                wf_cfg["run_until_stopped"] = False
                wf_cfg["run_until_queue_empty"] = True

            selectors = self.get_effective_selectors(workflow_cfg)
            print(f"\nStarting workflow: {workflow_name}")
            if bool(wf_cfg.get("process_invoice_list", False)):
                self.process_invoice_list(
                    page=page,
                    workflow_name=workflow_name,
                    workflow_cfg=wf_cfg,
                    selectors=selectors,
                    do_submit=do_submit,
                )
            elif bool(wf_cfg.get("process_in_list", False)):
                self.process_orders_in_list(
                    page=page,
                    workflow_name=workflow_name,
                    workflow_cfg=wf_cfg,
                    selectors=selectors,
                    do_submit=do_submit,
                )
            else:
                links = self.get_order_links(page, workflow_name, wf_cfg, selectors)
                for order_url in links:
                    try:
                        self.process_order(
                            page,
                            order_url,
                            selectors=selectors,
                            workflow_name=workflow_name,
                            workflow_cfg=wf_cfg,
                            do_submit=do_submit,
                        )
                    except Exception as ex:
                        self.stats["orders_failed"] += 1
                        print(f"[{workflow_name}] Failed processing order {order_url}: {ex}")

    def run(self, do_submit: bool, workflow_filter: str) -> None:
        self.load_csv_index()
        with sync_playwright() as playwright:
            browser = self.launch_browser(playwright)
            context = browser.new_context()
            page = context.new_page()

            try:
                self.login(page)
                self.run_workflows_after_login(page, do_submit=do_submit, workflow_filter=workflow_filter)
            finally:
                if not do_submit:
                    auto = self.config.get("automation", {})
                    if workflow_filter == "ship_to_store":
                        hold_sec = int(auto.get("dry_run_browser_hold_seconds_ship_to_store", 20))
                    elif workflow_filter == "ship_to_customer":
                        hold_sec = int(auto.get("dry_run_browser_hold_seconds_ship_to_customer", 10))
                    elif workflow_filter == "invoice":
                        hold_sec = int(auto.get("dry_run_browser_hold_seconds_invoice", 15))
                    elif workflow_filter == "all":
                        hold_sec = int(auto.get("dry_run_browser_hold_seconds", 15))
                    else:
                        hold_sec = int(auto.get("dry_run_browser_hold_seconds", 15))
                    if hold_sec > 0:
                        print(
                            f"\nDry run: holding browser open for {hold_sec}s so you can verify fields..."
                        )
                        try:
                            page.wait_for_timeout(hold_sec * 1000)
                        except Exception:
                            pass
                context.close()
                browser.close()

        print("\nAutomation complete.")
        print(json.dumps(self.stats, indent=2))


def validate_config(config: Dict) -> None:
    required_paths = [
        ["rithum", "login_url"],
        ["rithum", "orders_url"],
        ["rithum", "username"],
        ["rithum", "password"],
        ["rithum", "selectors", "username_input"],
        ["rithum", "selectors", "password_input"],
        ["rithum", "selectors", "login_button"],
        ["rithum", "selectors", "logged_in_ready"],
        ["rithum", "selectors", "order_links"],
        ["rithum", "selectors", "po_on_order_page"],
        ["rithum", "selectors", "tracking_input"],
        ["rithum", "selectors", "shipment_type_input"],
        ["rithum", "selectors", "quantity_input"],
        ["rithum", "selectors", "submit_button"],
        ["fedex_csv", "shipment_type_value"],
    ]
    for path in required_paths:
        require_string(config, path)

    fedex = config["fedex_csv"]
    has_direct_path = isinstance(fedex.get("path"), str) and fedex.get("path", "").strip()
    has_directory_and_base = (
        isinstance(fedex.get("directory"), str)
        and fedex.get("directory", "").strip()
        and isinstance(fedex.get("base_name"), str)
        and fedex.get("base_name", "").strip()
    )
    if not has_direct_path and not has_directory_and_base:
        raise ValueError(
            "fedex_csv must include either a file 'path' or both 'directory' and 'base_name'"
        )

    has_letter_mapping = (
        isinstance(fedex.get("po_column_letter"), str)
        and fedex.get("po_column_letter", "").strip()
        and isinstance(fedex.get("tracking_column_letter"), str)
        and fedex.get("tracking_column_letter", "").strip()
    )
    has_header_mapping = (
        isinstance(fedex.get("po_column"), str)
        and fedex.get("po_column", "").strip()
        and isinstance(fedex.get("tracking_column"), str)
        and fedex.get("tracking_column", "").strip()
    )
    if not has_letter_mapping and not has_header_mapping:
        raise ValueError(
            "fedex_csv must include either column letters "
            "('po_column_letter' and 'tracking_column_letter') "
            "or header names ('po_column' and 'tracking_column')"
        )

    workflows = config["rithum"].get("workflows")
    if workflows is not None:
        if not isinstance(workflows, dict) or not workflows:
            raise ValueError("rithum.workflows must be a non-empty object when provided")
        has_enabled = False
        for workflow_name, workflow_cfg in workflows.items():
            if not isinstance(workflow_cfg, dict):
                raise ValueError(f"Workflow '{workflow_name}' must be an object")
            if not bool(workflow_cfg.get("enabled", True)):
                continue
            has_enabled = True
            orders_url = workflow_cfg.get("orders_url")
            if not isinstance(orders_url, str) or not orders_url.strip():
                raise ValueError(f"Workflow '{workflow_name}' must define a non-empty orders_url")
        if not has_enabled:
            raise ValueError("At least one rithum workflow must be enabled")


def load_config(path: Path) -> Dict:
    with path.open("r", encoding="utf-8") as file:
        config = json.load(file)
    config = deep_expand_env(config)
    validate_config(config)
    return config


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Automate Lowe's tracking and invoicing in Rithum (CommerceHub) using FedEx file data."
    )
    parser.add_argument(
        "--config",
        default="config.example.json",
        help="Path to JSON config file. Default: config.example.json",
    )
    parser.add_argument(
        "--submit",
        action="store_true",
        help="Actually click Submit. Omit for dry run.",
    )
    parser.add_argument(
        "--workflow",
        default="all",
        help="Workflow to run: all, ship_to_store, ship_to_customer, invoice, or default",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config_path = Path(args.config)
    if not config_path.exists():
        print(
            f"Config file not found: {config_path}\n"
            "Provide a valid config file path (for example: config.example.json)."
        )
        return 1

    try:
        config = load_config(config_path)
        automation = LowesTrackingAutomation(config)
        automation.run(do_submit=args.submit, workflow_filter=args.workflow)
        return 0
    except TimeoutError as ex:
        print(f"Playwright timeout: {ex}")
        return 2
    except Exception as ex:
        print(f"Error: {ex}")
        return 3


if __name__ == "__main__":
    sys.exit(main())
