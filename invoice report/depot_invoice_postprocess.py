"""
Post-process CommerceHub invoice exports (Depot / Lowe's / Tractor): Accounting, PO line dedupe, totals, UNC xlsx, Excel print.

Same column layout for both retailers; output folder and filename prefix differ by retailer.
"""

from __future__ import annotations

import csv
import io
import os
import re
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Default shares (override in .env)
DEFAULT_DEPOT_OUTPUT_DIR = r"\\rygarcorp.com\shares\Cornerstone\Invoice Reports\Depot"
DEFAULT_LOWE_OUTPUT_DIR = r"\\rygarcorp.com\shares\Cornerstone\Invoice Reports\Lowe's"
DEFAULT_TRACTOR_OUTPUT_DIR = r"\\rygarcorp.com\shares\Cornerstone\Invoice Reports\Tractor Supply"

# Excel accounting format; backslashes must be literal (raw string avoids SyntaxWarning).
_DEFAULT_ACCOUNTING_NUMBER_FORMAT = (
    r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
)


def _accounting_format() -> str:
    return (
        os.environ.get(
            "COMMERCEHUB_ACCOUNTING_NUMBER_FORMAT",
            _DEFAULT_ACCOUNTING_NUMBER_FORMAT,
        ).strip()
        or _DEFAULT_ACCOUNTING_NUMBER_FORMAT
    )


def depot_report_filename(report_day: date) -> str:
    """e.g. Depot Invoice Report 5-12-2026 (no leading zeros on month/day)."""
    return f"Depot Invoice Report {report_day.month}-{report_day.day}-{report_day.year}.xlsx"


def lowes_report_filename(report_day: date) -> str:
    """e.g. Lowe's Invoice Report 5-12-2026."""
    return f"Lowe's Invoice Report {report_day.month}-{report_day.day}-{report_day.year}.xlsx"


def tractor_report_filename(report_day: date) -> str:
    """Tractor workbook name aligned with Depot/Lowe's date style (month-day-year, no leading zeros)."""
    return f"Tractor Supply Invoice Report {report_day.month}-{report_day.day}-{report_day.year}.xlsx"


def _excel_col_letters_to_zero_based(letters: str) -> int:
    """A→0, Z→25, AA→26, CA→78."""
    n = 0
    for ch in letters.upper().strip():
        if ch < "A" or ch > "Z":
            raise ValueError(f"Invalid Excel column letter: {letters!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


# Raw SPS columns → output A..G (consecutive). Column H is Retailer (appended).
# A,B,E,M,O,CA,R → output A..G (SKU / Vendor Style in G).
_TRACTOR_RAW_INDICES_OUT_ORDER: tuple[int, ...] = (
    _excel_col_letters_to_zero_based("A"),
    _excel_col_letters_to_zero_based("B"),
    _excel_col_letters_to_zero_based("E"),
    _excel_col_letters_to_zero_based("M"),
    _excel_col_letters_to_zero_based("O"),
    _excel_col_letters_to_zero_based("CA"),
    _excel_col_letters_to_zero_based("R"),
)
_TRACTOR_RETAILER_HEADER = "Retailer"
_TRACTOR_RETAILER_VALUE = "Tractor Supply"
# 1-based sheet columns for Accounting / SUM (same convention as Depot workbook).
_TRACTOR_COL_UNIT_PRICE = 5  # E
_TRACTOR_COL_CA = 6  # F
_TRACTOR_COL_QTY = 4  # D — center alignment
# openpyxl column_dimensions letter → width (D width not specified; use Excel default).
_TRACTOR_COLUMN_WIDTHS: dict[str, float] = {
    "A": 14.50,
    "B": 11.5,
    "C": 11.5,
    "E": 10.2,
    "F": 14.5,
    "G": 19.0,
    "H": 16.2,
}


def _tractor_append_retailer_column(rows: list[list[str]]) -> list[list[str]]:
    """Column H: header ``Retailer``, each data row ``Tractor Supply``."""
    if not rows:
        return rows
    out: list[list[str]] = [rows[0] + [_TRACTOR_RETAILER_HEADER]]
    for row in rows[1:]:
        out.append(row + [_TRACTOR_RETAILER_VALUE])
    return out


def _tractor_coerce_numeric(val) -> float | str | None:
    """Parse currency-like strings for Excel Accounting; return None for blank."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    cleaned = s.replace(",", "").replace("$", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return val


def _tractor_norm_header_row(row: list[str]) -> tuple[str, ...]:
    return tuple((c or "").strip().lower() for c in row)


def _tractor_read_plain_csv_file(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        return list(csv.reader(f))


def _tractor_csv_rows_from_download(path: Path) -> list[list[str]]:
    """
    SPS often delivers a ZIP of per-invoice CSVs (still named ``.csv`` on disk).
    Plain CSV is returned as-is; ZIP members ending in ``.csv`` are merged (one header).
    """
    if zipfile.is_zipfile(path):
        merged: list[list[str]] = []
        header_key: tuple[str, ...] | None = None
        with zipfile.ZipFile(path, "r") as zf:
            names = sorted(
                n
                for n in zf.namelist()
                if n.lower().endswith(".csv") and not n.replace("\\", "/").startswith("__")
            )
            if not names:
                raise RuntimeError(
                    f"Tractor SPS download is a ZIP but contains no .csv entries: {path}"
                )
            for name in names:
                info = zf.getinfo(name)
                if info.is_dir():
                    continue
                with zf.open(name, "r") as bf:
                    text = io.TextIOWrapper(
                        bf, encoding="utf-8-sig", errors="replace", newline=""
                    )
                    part = list(csv.reader(text))
                if not part:
                    continue
                if header_key is None:
                    header_key = _tractor_norm_header_row(part[0])
                    merged.extend(part)
                    continue
                start = 1 if _tractor_norm_header_row(part[0]) == header_key else 0
                merged.extend(part[start:])
        if not merged:
            raise RuntimeError(f"No CSV rows could be read from ZIP: {path}")
        return merged
    return _tractor_read_plain_csv_file(path)


def _tractor_project_report_columns(rows: list[list[str]]) -> list[list[str]]:
    """Map raw A,B,E,M,O,CA,R to output columns A..G (0-based row lists length 7)."""
    out: list[list[str]] = []
    for row in rows:
        out.append([row[i] if i < len(row) else "" for i in _TRACTOR_RAW_INDICES_OUT_ORDER])
    return out


def _tractor_drop_stub_body_rows(rows: list[list[str]]) -> list[list[str]]:
    """
    Raw exports repeat the invoice # on a following row with no other fields; drop those
    after projection (columns A..G) so column A stays line-item rows only (plus header).
    """
    if not rows:
        return rows
    out = [rows[0]]
    for row in rows[1:]:
        if not any((c or "").strip() for c in row):
            continue
        if (row[0] or "").strip() and not any((row[j] or "").strip() for j in range(1, len(row))):
            continue
        out.append(row)
    return out


# 0-based output column indices (A..H after retailer append).
_TRACTOR_COL_INV = 0
_TRACTOR_COL_INV_DATE = 1
_TRACTOR_COL_PO_DATE = 2
_TRACTOR_COL_QTY_IX = 3
_TRACTOR_COL_UNIT_IX = 4
_TRACTOR_COL_TOTAL_IX = 5
_TRACTOR_COL_STYLE_IX = 6
_TRACTOR_HEADER_COLS: tuple[int, ...] = (
    _TRACTOR_COL_INV_DATE,
    _TRACTOR_COL_PO_DATE,
    _TRACTOR_COL_TOTAL_IX,
)
_TRACTOR_DETAIL_COLS: tuple[int, ...] = (
    _TRACTOR_COL_QTY_IX,
    _TRACTOR_COL_UNIT_IX,
    _TRACTOR_COL_STYLE_IX,
)
_TRACTOR_REQUIRED_FIELDS: tuple[tuple[int, str], ...] = (
    (_TRACTOR_COL_INV, "Invoice Number"),
    (_TRACTOR_COL_INV_DATE, "Invoice Date"),
    (_TRACTOR_COL_PO_DATE, "PO Date"),
    (_TRACTOR_COL_QTY_IX, "Qty Ordered"),
    (_TRACTOR_COL_UNIT_IX, "Unit Price"),
    (_TRACTOR_COL_TOTAL_IX, "Invoice Total"),
    (_TRACTOR_COL_STYLE_IX, "Vendor Style"),
)
_TRACTOR_AMOUNT_TOLERANCE = 0.02


def _tractor_note(notes: list[str], message: str) -> None:
    notes.append(message)


def _tractor_cell_text(row: list[str], col_ix: int) -> str:
    if col_ix >= len(row):
        return ""
    return (row[col_ix] or "").strip()


def _tractor_pad_row(row: list[str], width: int) -> list[str]:
    padded = list(row)
    if len(padded) < width:
        padded.extend([""] * (width - len(padded)))
    return padded


def _tractor_row_has_header_fields(row: list[str]) -> bool:
    return any(_tractor_cell_text(row, col_ix) for col_ix in _TRACTOR_HEADER_COLS)


def _tractor_row_has_detail_fields(row: list[str]) -> bool:
    return any(_tractor_cell_text(row, col_ix) for col_ix in _TRACTOR_DETAIL_COLS)


def _tractor_row_is_header_only(row: list[str]) -> bool:
    """Header half of a split line: dates and/or line total, no SKU/qty/price."""
    if _tractor_cell_text(row, _TRACTOR_COL_STYLE_IX):
        return False
    if _tractor_cell_text(row, _TRACTOR_COL_QTY_IX) or _tractor_cell_text(row, _TRACTOR_COL_UNIT_IX):
        return False
    return _tractor_row_has_header_fields(row)


def _tractor_row_is_detail_only(row: list[str]) -> bool:
    """Detail half of a split line: SKU and/or qty/price, no header dates/total."""
    if _tractor_cell_text(row, _TRACTOR_COL_STYLE_IX):
        return True
    has_qty_or_price = bool(
        _tractor_cell_text(row, _TRACTOR_COL_QTY_IX) or _tractor_cell_text(row, _TRACTOR_COL_UNIT_IX)
    )
    return has_qty_or_price and not _tractor_row_has_header_fields(row)


def _tractor_first_nonempty(block: list[list[str]], col_ix: int, *, rows_first: list[list[str]]) -> str:
    for r in rows_first + block:
        val = _tractor_cell_text(r, col_ix)
        if val:
            return val
    return ""


def _tractor_line_label(inv: str, sku: str) -> str:
    return f"{inv} SKU {sku}" if sku else inv


def _tractor_validate_merged_row(merged: list[str], label: str, notes: list[str]) -> None:
    missing = [
        label for col_ix, label in _TRACTOR_REQUIRED_FIELDS if not _tractor_cell_text(merged, col_ix)
    ]
    if missing:
        _tractor_note(
            notes,
            f"Tractor {label}: incomplete row after merge (missing: {', '.join(missing)}).",
        )

    qty = _tractor_coerce_numeric(_tractor_cell_text(merged, _TRACTOR_COL_QTY_IX))
    unit = _tractor_coerce_numeric(_tractor_cell_text(merged, _TRACTOR_COL_UNIT_IX))
    total = _tractor_coerce_numeric(_tractor_cell_text(merged, _TRACTOR_COL_TOTAL_IX))
    if isinstance(qty, (int, float)) and isinstance(unit, (int, float)) and isinstance(total, (int, float)):
        expected = float(qty) * float(unit)
        if abs(expected - float(total)) > _TRACTOR_AMOUNT_TOLERANCE:
            _tractor_note(
                notes,
                f"Tractor {label}: qty ({qty}) × unit price ({unit}) = {expected:.2f} "
                f"but invoice total is {float(total):.2f}.",
            )


def _tractor_pick_detail_row(group: list[list[str]]) -> list[str] | None:
    for r in group:
        if _tractor_cell_text(r, _TRACTOR_COL_STYLE_IX):
            return r
    for r in group:
        if _tractor_row_is_detail_only(r):
            return r
    return None


def _tractor_split_po_block_into_line_groups(
    block: list[list[str]], inv: str, notes: list[str]
) -> list[list[list[str]]]:
    """
    Within one PO (same invoice #), SPS emits a header row + detail row per SKU.
    Pair them so each SKU becomes its own merge group (2 raw rows → 1 output row).
    """
    if len(block) == 1:
        return [block]

    groups: list[list[list[str]]] = []
    pending_header: list[str] | None = None
    i = 0
    while i < len(block):
        r = block[i]
        if _tractor_row_is_header_only(r):
            if pending_header is not None:
                _tractor_note(
                    notes,
                    f"Tractor invoice {inv}: back-to-back header rows — using nearest header per line.",
                )
            pending_header = r
            i += 1
            continue
        if _tractor_row_is_detail_only(r):
            group: list[list[str]] = []
            if pending_header is not None:
                group.append(pending_header)
                pending_header = None
            group.append(r)
            groups.append(group)
            i += 1
            continue
        if pending_header is not None:
            groups.append([pending_header, r])
            pending_header = None
        else:
            groups.append([r])
        i += 1
    if pending_header is not None:
        _tractor_note(
            notes,
            f"Tractor invoice {inv}: header row without a matching line item — row kept as-is.",
        )
        groups.append([pending_header])
    return groups if groups else [block]


def _tractor_merge_line_group(
    group: list[list[str]],
    block: list[list[str]],
    inv: str,
    width: int,
    notes: list[str],
) -> list[str]:
    """Merge one header+detail pair (or a single complete row) into one output line."""
    if len(group) == 1 and not _tractor_row_is_header_only(group[0]):
        merged = _tractor_pad_row(group[0], width)
        sku = _tractor_cell_text(merged, _TRACTOR_COL_STYLE_IX)
        _tractor_validate_merged_row(merged, _tractor_line_label(inv, sku), notes)
        return merged

    header_rows = [r for r in group if _tractor_row_has_header_fields(r)]
    block_headers = [r for r in block if _tractor_row_has_header_fields(r)]
    detail_row = _tractor_pick_detail_row(group)

    merged = [""] * width
    merged[_TRACTOR_COL_INV] = inv
    merged[_TRACTOR_COL_INV_DATE] = (
        _tractor_first_nonempty(group, _TRACTOR_COL_INV_DATE, rows_first=header_rows)
        or _tractor_first_nonempty(block, _TRACTOR_COL_INV_DATE, rows_first=block_headers)
    )
    merged[_TRACTOR_COL_PO_DATE] = (
        _tractor_first_nonempty(group, _TRACTOR_COL_PO_DATE, rows_first=header_rows)
        or _tractor_first_nonempty(block, _TRACTOR_COL_PO_DATE, rows_first=block_headers)
    )
    merged[_TRACTOR_COL_TOTAL_IX] = _tractor_first_nonempty(
        group, _TRACTOR_COL_TOTAL_IX, rows_first=header_rows
    )
    if detail_row is not None:
        merged[_TRACTOR_COL_QTY_IX] = _tractor_cell_text(detail_row, _TRACTOR_COL_QTY_IX)
        merged[_TRACTOR_COL_UNIT_IX] = _tractor_cell_text(detail_row, _TRACTOR_COL_UNIT_IX)
        merged[_TRACTOR_COL_STYLE_IX] = _tractor_cell_text(detail_row, _TRACTOR_COL_STYLE_IX)
    if len(merged) > 7:
        merged[7] = (
            _tractor_first_nonempty(group, 7, rows_first=group) or _TRACTOR_RETAILER_VALUE
        )

    sku = _tractor_cell_text(merged, _TRACTOR_COL_STYLE_IX)
    _tractor_validate_merged_row(merged, _tractor_line_label(inv, sku), notes)
    return merged


def _tractor_consolidate_po_rows_per_invoice(
    rows: list[list[str]], notes: list[str]
) -> list[list[str]]:
    """
    SPS splits each PO line across two rows (header + SKU detail). Within each invoice # block,
    pair header/detail rows per SKU so one SKU → one output row (two SKUs → two output rows).
    """
    if len(rows) <= 1:
        return rows
    width = max(len(r) for r in rows)
    out: list[list[str]] = [_tractor_pad_row(rows[0], width)]
    seen_po_sku: set[tuple[str, str]] = set()
    i = 1
    while i < len(rows):
        inv = _tractor_cell_text(rows[i], _TRACTOR_COL_INV)
        if not inv:
            out.append(_tractor_pad_row(rows[i], width))
            i += 1
            continue
        block: list[list[str]] = []
        while i < len(rows) and _tractor_cell_text(rows[i], _TRACTOR_COL_INV) == inv:
            block.append(_tractor_pad_row(rows[i], width))
            i += 1
        for group in _tractor_split_po_block_into_line_groups(block, inv, notes):
            merged = _tractor_merge_line_group(group, block, inv, width, notes)
            sku = _tractor_cell_text(merged, _TRACTOR_COL_STYLE_IX)
            key = (inv, sku)
            if key in seen_po_sku:
                _tractor_note(
                    notes,
                    f"Tractor report: duplicate PO {inv} SKU {sku or '(blank)'} — kept first row only.",
                )
                continue
            seen_po_sku.add(key)
            out.append(merged)
    return out


def save_tractor_supply_csv(source: Path, report_day: date) -> tuple[Path, list[str]]:
    """
    Read the SPS download (plain CSV or a ZIP of CSVs), project raw columns A,B,E,M,O,CA,R to
    sheet columns A..G, append ``Retailer`` / ``Tractor Supply`` in column H, drop invoice-only
    stub lines, merge split header/detail rows into one row per PO line (invoice # + SKU), apply Accounting to
    columns E and F (Depot/Lowe's style), set column widths and center column D, total column F
    with a SUM row (total cell Accounting), and save ``.xlsx`` to the Tractor Supply share.

    Returns ``(workbook_path, review_notes)``. Notes list anything worth checking before
    QuickBooks entry; the report is always saved when rows exist.
    Uses COMMERCEHUB_TRACTOR_OUTPUT_DIR and COMMERCEHUB_OUTPUT_FALLBACK_DIR like Depot workbooks.
    """
    from dotenv import load_dotenv

    load_dotenv()
    review_notes: list[str] = []
    raw_rows = _tractor_csv_rows_from_download(source)
    rows = _tractor_consolidate_po_rows_per_invoice(
        _tractor_append_retailer_column(
            _tractor_drop_stub_body_rows(_tractor_project_report_columns(raw_rows))
        ),
        review_notes,
    )
    if not rows:
        raise RuntimeError("Tractor SPS export produced no rows after download merge.")

    out_dir = Path(os.environ.get("COMMERCEHUB_TRACTOR_OUTPUT_DIR", DEFAULT_TRACTOR_OUTPUT_DIR))
    out_path = out_dir / tractor_report_filename(report_day)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Report"
    header_row = 1
    first_data = 2

    for c, text in enumerate(rows[0], start=1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.font = Font(bold=True)

    for ri, row in enumerate(rows[1:], start=first_data):
        for ci, val in enumerate(row, start=1):
            if ci == _TRACTOR_COL_QTY:
                num = _tractor_coerce_numeric(val)
                if isinstance(num, (int, float)):
                    ws.cell(row=ri, column=ci, value=int(num) if num == int(num) else num)
                elif not str(val).strip():
                    ws.cell(row=ri, column=ci, value=None)
                else:
                    ws.cell(row=ri, column=ci, value=val)
            elif ci == _TRACTOR_COL_UNIT_PRICE or ci == _TRACTOR_COL_CA:
                num = _tractor_coerce_numeric(val)
                if num is not None:
                    ws.cell(row=ri, column=ci, value=num)
                elif not str(val).strip():
                    ws.cell(row=ri, column=ci, value=None)
                else:
                    ws.cell(row=ri, column=ci, value=val)
            else:
                ws.cell(row=ri, column=ci, value=val)

    acct = _accounting_format()
    last_data = header_row + len(rows) - 1
    sum_row: int | None = None
    if last_data >= first_data:
        sum_row = last_data + 1
        lf = get_column_letter(_TRACTOR_COL_CA)
        ws.cell(
            row=sum_row,
            column=_TRACTOR_COL_CA,
            value=f"=SUM({lf}{first_data}:{lf}{last_data})",
        )
        for ri in range(first_data, last_data + 1):
            ws.cell(row=ri, column=_TRACTOR_COL_UNIT_PRICE).number_format = acct
        for ri in range(header_row, sum_row + 1):
            ws.cell(row=ri, column=_TRACTOR_COL_CA).number_format = acct
        ws.cell(row=sum_row, column=_TRACTOR_COL_CA).number_format = acct
    else:
        for ri in range(first_data, last_data + 1):
            ws.cell(row=ri, column=_TRACTOR_COL_UNIT_PRICE).number_format = acct
        for ri in range(header_row, last_data + 1):
            ws.cell(row=ri, column=_TRACTOR_COL_CA).number_format = acct

    end_layout_row = sum_row if sum_row is not None else last_data
    center = Alignment(horizontal="center", vertical="center")
    for letter, width in _TRACTOR_COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for ri in range(header_row, end_layout_row + 1):
        ws.cell(row=ri, column=_TRACTOR_COL_QTY).alignment = center

    _apply_openpyxl_print_file_and_page_footer(ws)

    return _save_xlsx_or_fallback(wb, out_path), review_notes


def parse_order_line(value) -> int | None:
    """Parse CommerceHub cells like =\"1\" or =1 into an integer."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    if not s:
        return None
    for pat in (
        r'^="(\d+)"\s*$',
        r"^='(\d+)'\s*$",
        r"^=(\d+)\s*$",
        r"^(\d+)\s*$",
    ):
        m = re.match(pat, s, re.I)
        if m:
            return int(m.group(1))
    return None


class InvoiceExportEmpty(Exception):
    """CommerceHub export has no invoice table (empty search day or summary-only file)."""


def _line_looks_like_invoice_header(line: str) -> bool:
    low = (line or "").lower()
    if "order line number" in low:
        return True
    if "order" in low and "line" in low and ("po" in low or "vendor" in low or "sku" in low):
        return True
    parts = re.split(r"[,\t;]", line)
    joined = " ".join(p.strip().lower() for p in parts if p.strip())
    if "order" in joined and "line" in joined and ("po" in joined or "vendor" in joined):
        return True
    return False


def _find_csv_header_index(lines: list[str]) -> int | None:
    for i, line in enumerate(lines[:60]):
        if _line_looks_like_invoice_header(line):
            return i
    return None


def _text_indicates_no_invoices(text: str) -> bool:
    low = (text or "").lower()
    needles = (
        "no record",
        "no results",
        "0 records",
        "no orders found",
        "search returned no",
        "did not match",
        "no matching",
        "there are no",
    )
    return any(n in low for n in needles)


def _find_column(columns: list[str], *must_contain: str, must_not_contain: tuple[str, ...] = ()) -> str:
    cols = [str(c).strip() for c in columns]
    for c in cols:
        u = c.upper()
        if all(k.upper() in u for k in must_contain) and all(x.upper() not in u for x in must_not_contain):
            return c
    raise KeyError(f"No column matching {must_contain!r} exclude {must_not_contain!r} in {cols!r}")


def read_invoice_export(path: Path) -> tuple[list[str], pd.DataFrame]:
    """Split metadata lines (above table) from the tabular data."""
    raw = path.read_bytes()
    if raw[:2] == b"PK":
        return _read_invoice_xlsx(path)

    suf = path.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return _read_invoice_xlsx(path)
    if suf == ".xls":
        raise RuntimeError(
            "Legacy .xls is not supported for post-process; re-export as CSV or xlsx, "
            "or install xlrd and extend depot_invoice_postprocess."
        )

    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RuntimeError(f"Could not decode export file: {path}")

    if "<html" in text[:500].lower():
        raise RuntimeError(
            f"Export file looks like HTML, not a CommerceHub invoice CSV: {path}"
        )

    lines = [ln for ln in text.splitlines() if ln is not None]
    header_i = _find_csv_header_index(lines)
    if header_i is None:
        if _text_indicates_no_invoices(text) or len(lines) < 12:
            raise InvoiceExportEmpty(
                f"No invoice rows in export (empty day or summary-only): {path}"
            )
        preview = "\n".join(lines[:6])
        raise RuntimeError(
            f"Could not find header row (Order Line / PO / Vendor) in {path}. "
            f"Check the export format. First lines:\n{preview}"
        )

    meta = lines[:header_i]
    body = "\n".join(lines[header_i:])
    try:
        dialect = csv.Sniffer().sniff(body[:8192], delimiters=",\t;")
        sep = dialect.delimiter
    except csv.Error:
        sep = ","
    df = pd.read_csv(io.StringIO(body), sep=sep, header=0, dtype=str, engine="python")
    df = df.dropna(axis=1, how="all")
    return meta, df


def _read_invoice_xlsx(path: Path) -> tuple[list[str], pd.DataFrame]:
    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = pd.read_excel(xl, sheet_name=0, header=None, dtype=str)
    header_row = None
    for i in range(min(30, len(sheet))):
        row = sheet.iloc[i].astype(str).str.lower().fillna("")
        joined = " ".join(row.tolist())
        if "order" in joined and "line" in joined and ("po" in joined or "vendor" in joined):
            header_row = i
            break
    if header_row is None:
        raise RuntimeError(f"Could not find header row in xlsx: {path}")
    meta_lines = []
    for j in range(header_row):
        parts = [str(x) for x in sheet.iloc[j].tolist() if str(x) != "nan" and str(x).strip()]
        meta_lines.append("\t".join(parts) if parts else "")
    headers = [str(x).strip() for x in sheet.iloc[header_row].tolist()]
    data = sheet.iloc[header_row + 1 :].copy()
    data.columns = headers[: len(data.columns)]
    data = data.dropna(axis=1, how="all")
    return meta_lines, data


def dedupe_po_lines(df: pd.DataFrame) -> pd.DataFrame:
    """Column A: 1 stays 1; line > 1 -> A=0 and F=0 for duplicate PO invoice totals."""
    df = df.copy()
    # "ORDER LINE NUMBER (ORDER LINE)" — avoid INVOICE/VENDOR/QUANTITY "(... LINE)" columns.
    line_col = _find_column(
        list(df.columns),
        "ORDER",
        "LINE",
        "NUMBER",
        must_not_contain=("INVOICE", "VENDOR", "QUANTITY"),
    )
    # "PO NUMBER (ORDER)" — not the order-line column.
    po_col = _find_column(list(df.columns), "PO", "NUMBER", must_not_contain=("LINE",))
    total_col = _find_column(list(df.columns), "INVOICE", "TOTAL", must_not_contain=("UNIT",))

    for po in df[po_col].unique():
        idx = df.index[df[po_col] == po].tolist()
        for ix in idx:
            n = parse_order_line(df.at[ix, line_col])
            if n is None:
                continue
            if n == 1:
                df.at[ix, line_col] = "1"
            else:
                df.at[ix, line_col] = "0"
                df.at[ix, total_col] = "0"

    df[line_col] = pd.to_numeric(df[line_col], errors="coerce").fillna(0).astype(int)
    df[total_col] = pd.to_numeric(
        df[total_col].astype(str).str.replace(",", "", regex=False).str.replace("$", "", regex=False),
        errors="coerce",
    ).fillna(0.0)
    return df


def _col_index_one_based(df: pd.DataFrame, col_name: str) -> int:
    return list(df.columns).index(col_name) + 1


def _apply_openpyxl_print_file_and_page_footer(ws) -> None:
    """
    Embed print header/footer in the saved workbook (Excel OOXML codes).
    Center header: file name; center footer: ``Page 1 of 4`` style pagination.
    """
    header_txt = "&[File]"
    footer_txt = "Page &[Page] of &[Pages]"
    for name in ("oddHeader", "evenHeader"):
        getattr(ws, name).center.text = header_txt
    for name in ("oddFooter", "evenFooter"):
        getattr(ws, name).center.text = footer_txt


def write_depot_workbook(
    meta_lines: list[str],
    df: pd.DataFrame,
    out_path: Path,
    *,
    unit_col: str,
    total_col: str,
) -> Path:
    """Write metadata + table + SUM rows; Accounting format on unit and total columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    r = 1
    for line in meta_lines:
        ws.cell(row=r, column=1, value=line)
        r += 1

    header_row = r
    for c, name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c, value=name)
        ws.cell(row=header_row, column=c).font = Font(bold=True)

    first_data_row = header_row + 1
    for ri, row in enumerate(df.itertuples(index=False), start=first_data_row):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    last_data_row = header_row + len(df)
    sum_row = last_data_row + 1

    line_col = _find_column(
        list(df.columns),
        "ORDER",
        "LINE",
        "NUMBER",
        must_not_contain=("INVOICE", "VENDOR", "QUANTITY"),
    )
    col_a = _col_index_one_based(df, line_col)
    col_f = _col_index_one_based(df, total_col)
    la = get_column_letter(col_a)
    lf = get_column_letter(col_f)

    ws.cell(row=sum_row, column=col_a, value=f"=SUM({la}{first_data_row}:{la}{last_data_row})")
    ws.cell(row=sum_row, column=col_f, value=f"=SUM({lf}{first_data_row}:{lf}{last_data_row})")

    col_e = _col_index_one_based(df, unit_col)
    col_f_idx = _col_index_one_based(df, total_col)
    acct = _accounting_format()
    for ri in range(first_data_row, last_data_row + 1):
        ws.cell(row=ri, column=col_e).number_format = acct

    # Entire invoice-total column (header through totals row) uses Accounting.
    for ri in range(header_row, sum_row + 1):
        ws.cell(row=ri, column=col_f_idx).number_format = acct

    ws.cell(row=sum_row, column=col_a).number_format = "0"

    # Column layout (Excel letters match standard Depot export: A=line … H=merchant).
    ws.column_dimensions["B"].width = 10.5
    ws.column_dimensions["C"].width = 19.0
    ws.column_dimensions["F"].width = 15.5
    ws.column_dimensions["G"].width = 24.0
    center = Alignment(horizontal="center", vertical="center")
    for ri in range(header_row, sum_row + 1):
        ws.cell(row=ri, column=4).alignment = center

    # Centering runs after formats; re-apply Accounting on the SUM cell (openpyxl edge cases).
    ws.cell(row=sum_row, column=col_f_idx).number_format = acct

    _apply_openpyxl_print_file_and_page_footer(ws)

    return _save_xlsx_or_fallback(wb, out_path)


def _save_xlsx_or_fallback(wb, out_path: Path) -> Path:
    """
    Save to the primary UNC/local path; on permission denied, optionally save under
    COMMERCEHUB_OUTPUT_FALLBACK_DIR so the run can still complete (e.g. file open on share).
    """
    out_path = out_path.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
        return out_path
    except OSError as exc:
        if not isinstance(exc, PermissionError) and getattr(exc, "errno", None) not in (13, 1):
            raise
        fb_raw = (os.environ.get("COMMERCEHUB_OUTPUT_FALLBACK_DIR") or "").strip()
        if not fb_raw:
            raise RuntimeError(
                f"Permission denied saving {out_path}. "
                "Close that workbook in Excel if it is open, confirm this Windows user can write to the share "
                "(mapped drive vs UNC can differ for scheduled tasks), or set COMMERCEHUB_OUTPUT_FALLBACK_DIR "
                "to a local folder to save a copy there automatically."
            ) from exc
        alt = Path(fb_raw).expanduser().resolve() / out_path.name
        alt.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.save(alt)
        except OSError as exc2:
            raise RuntimeError(
                f"Could not save to primary ({out_path}) or fallback ({alt}): {exc2}"
            ) from exc2
        print(f"[invoice] Share not writable; saved to fallback: {alt}", flush=True)
        return alt


def process_invoice_download(downloaded: Path, report_day: date, retailer: str) -> Path:
    """
    retailer: \"depot\" | \"lowes\" — same spreadsheet rules; different output folder and filename.
    """
    from dotenv import load_dotenv

    load_dotenv()

    r = retailer.strip().lower()
    if r == "depot":
        if (os.environ.get("COMMERCEHUB_DEPOT_POSTPROCESS") or "true").strip().lower() in (
            "0",
            "false",
            "no",
        ):
            return downloaded
        out_dir = Path(os.environ.get("COMMERCEHUB_DEPOT_OUTPUT_DIR", DEFAULT_DEPOT_OUTPUT_DIR))
        out_name = depot_report_filename(report_day)
    elif r == "lowes":
        if (os.environ.get("COMMERCEHUB_LOWE_POSTPROCESS") or "true").strip().lower() in (
            "0",
            "false",
            "no",
        ):
            return downloaded
        out_dir = Path(os.environ.get("COMMERCEHUB_LOWE_OUTPUT_DIR", DEFAULT_LOWE_OUTPUT_DIR))
        out_name = lowes_report_filename(report_day)
    else:
        raise ValueError(f"Unknown retailer: {retailer!r} (use depot or lowes)")

    out_path = out_dir / out_name

    meta, df = read_invoice_export(downloaded)
    df = df.dropna(how="all")
    po_key = _find_column(list(df.columns), "PO", "NUMBER", must_not_contain=("LINE",))
    df = df[df[po_key].astype(str).str.strip() != ""]

    unit_col = _find_column(list(df.columns), "INVOICE", "UNIT")
    total_col = _find_column(list(df.columns), "INVOICE", "TOTAL", must_not_contain=("UNIT",))

    df2 = dedupe_po_lines(df)
    df2[unit_col] = pd.to_numeric(
        df2[unit_col].astype(str).str.replace(",", "", regex=False).str.replace("$", "", regex=False),
        errors="coerce",
    ).fillna(0.0)

    saved = write_depot_workbook(meta, df2, out_path, unit_col=unit_col, total_col=total_col)

    try:
        from depot_excel_print import print_landscape_with_gridlines

        print_landscape_with_gridlines(saved)
    except Exception as e:
        import traceback

        print(f"[invoice:{r}] Excel print step failed (workbook was saved): {e}", flush=True)
        traceback.print_exc()

    return saved


def process_depot_download(downloaded: Path, report_day: date) -> Path:
    """Backward-compatible wrapper for Depot-only callers."""
    return process_invoice_download(downloaded, report_day, "depot")


def process_lowes_download(downloaded: Path, report_day: date) -> Path:
    return process_invoice_download(downloaded, report_day, "lowes")
