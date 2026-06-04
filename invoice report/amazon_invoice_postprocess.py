"""
Post-process Amazon transaction exports from the Cornerstone share: trim header junk,
filter by run-day rules, keep Order rows only, forward-fill blank PO/SKU on continuation lines,
merge duplicate PO+SKU lines (sum qty and sales),
format columns, save, and print (landscape + gridlines).

Drop a new raw .csv/.xlsx into the Amazon **Input** share folder. Formatted output is saved to
**Output** as ``{same base name} Output.xlsx`` (then printed).

Date rules (by run date = today unless overridden) — **calendar date only**; time of day is ignored:
  Tue–Fri: keep Order rows whose date/time column is **yesterday's calendar date** (any time).
  Monday: keep Order rows on the previous **Friday, Saturday, and Sunday** (any time).

  Optional (default on): also include the **prior settlement calendar day** (Tue–Fri only).
  Amazon often labels many of yesterday's sales with the previous date in the export; disable with
  ``AMAZON_INCLUDE_PRIOR_SETTLEMENT_DAY=false`` if you only want the strict yesterday date.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from depot_invoice_postprocess import (
    _accounting_format,
    _apply_openpyxl_print_file_and_page_footer,
    _save_xlsx_or_fallback,
)

_SCRIPT_DIR = Path(__file__).resolve().parent
_ENV_FILE = _SCRIPT_DIR / ".env"
DEFAULT_AMAZON_BASE_DIR = r"\\rygarcorp.com\shares\Cornerstone\Invoice Reports\Amazon"
DEFAULT_AMAZON_INPUT_DIR = DEFAULT_AMAZON_BASE_DIR + r"\Input"
DEFAULT_AMAZON_OUTPUT_DIR = DEFAULT_AMAZON_BASE_DIR + r"\Output"

# Default raw column letters → output A..F (overridden per file from header names).
_DEFAULT_RAW_COL_INDICES: tuple[int, ...] = (0, 2, 3, 4, 6, 13)  # A, C, D, E, G, N

_AMAZON_TXN_RE = re.compile(
    r"^(?P<dt>[A-Za-z]{3}\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}:\d{2}\s+[AP]M)"
    r"(?:\s+[A-Z]{2,5})?",
    re.IGNORECASE,
)
_OUTPUT_STEM_SUFFIX = " Output"
_ELIGIBLE_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
_SKIP_NAME_PREFIXES = ("~$", ".")


def load_project_dotenv() -> None:
    load_dotenv(_ENV_FILE)


def resolve_amazon_base_dir() -> Path:
    raw = (os.environ.get("AMAZON_INVOICE_BASE_DIR") or DEFAULT_AMAZON_BASE_DIR).strip()
    return Path(raw).expanduser()


def resolve_amazon_input_dir() -> Path:
    raw = (os.environ.get("AMAZON_INVOICE_INPUT_DIR") or DEFAULT_AMAZON_INPUT_DIR).strip()
    return Path(raw).expanduser()


def resolve_amazon_output_dir() -> Path:
    raw = (os.environ.get("AMAZON_INVOICE_OUTPUT_DIR") or DEFAULT_AMAZON_OUTPUT_DIR).strip()
    return Path(raw).expanduser()


def output_path_for_source(source: Path, output_dir: Path | None = None) -> Path:
    """e.g. ``report.csv`` -> ``report Output.xlsx`` in the Output folder."""
    out_dir = (output_dir or resolve_amazon_output_dir()).resolve()
    return out_dir / f"{source.stem}{_OUTPUT_STEM_SUFFIX}.xlsx"


def _include_prior_settlement_day() -> bool:
    """Default true — see module docstring."""
    v = (os.environ.get("AMAZON_INCLUDE_PRIOR_SETTLEMENT_DAY") or "true").strip().lower()
    return v not in ("0", "false", "no", "")


def transaction_dates_to_keep(run_day: date | None = None) -> set[date]:
    """
    Calendar dates to keep (time of day is not used).

    Tue–Fri: yesterday; also day-before-yesterday when ``AMAZON_INCLUDE_PRIOR_SETTLEMENT_DAY`` is true.
    Monday: previous Fri, Sat, and Sun.
    """
    d = run_day or date.today()
    if d.weekday() == 0:
        return {d - timedelta(days=3), d - timedelta(days=2), d - timedelta(days=1)}
    days = {d - timedelta(days=1)}
    if _include_prior_settlement_day():
        days.add(d - timedelta(days=2))
    return days


def parse_amazon_transaction_datetime(value: object) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    s = str(value).strip()
    if not s:
        return None
    m = _AMAZON_TXN_RE.match(s)
    if not m:
        return None
    try:
        return datetime.strptime(m.group("dt"), "%b %d, %Y %I:%M:%S %p")
    except ValueError:
        return None


def parse_amazon_transaction_date(value: object) -> date | None:
    """Calendar date from the date/time cell; ignores time and timezone suffix."""
    dt = parse_amazon_transaction_datetime(value)
    return dt.date() if dt else None


def _row_matches_keep_dates(row: tuple, keep_dates: set[date]) -> bool:
    """True when the row's calendar date is in *keep_dates* (any time on that day counts)."""
    tx_date = parse_amazon_transaction_date(row[0] if row else None)
    return tx_date is not None and tx_date in keep_dates


def _row_width(row: tuple) -> int:
    for i in range(len(row) - 1, -1, -1):
        v = row[i]
        if v is not None and str(v).strip():
            return i + 1
    return 0


def _pad_row(row: tuple, width: int) -> tuple:
    w = max(width, _row_width(row))
    if len(row) >= w:
        return row
    return tuple(row) + (None,) * (w - len(row))


def _xlsx_cell_value(cell) -> object:
    """Prefer stable strings for order IDs (Excel often stores them as imprecise floats)."""
    val = cell.value
    if val is None:
        return None
    if cell.data_type == "s" or isinstance(val, str):
        return str(val).strip() if isinstance(val, str) else val
    if isinstance(val, float):
        if val == int(val) and abs(val) < 9_007_199_254_740_992:
            return str(int(val))
        return format(val, ".0f") if abs(val) >= 1e11 else val
    return val


def _load_rows(path: Path) -> list[tuple]:
    suf = path.suffix.lower()
    if suf == ".csv":
        import csv

        rows: list[tuple] = []
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
            for row in csv.reader(f):
                rows.append(tuple(row))
        return rows

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [tuple(_xlsx_cell_value(c) for c in row) for row in ws.iter_rows()]
    finally:
        wb.close()


def _header_cell_matches(cell: object, *parts: str, exclude: tuple[str, ...] = ()) -> bool:
    u = str(cell or "").strip().lower()
    if not u:
        return False
    return all(p in u for p in parts) and all(e not in u for e in exclude)


def _resolve_raw_col_indices(header: tuple) -> tuple[int, ...]:
    """Map Amazon export columns by header text (layout varies); keep output A..F shape."""
    width = max(29, _row_width(header))
    h = _pad_row(header, width)

    def find(*parts: str, exclude: tuple[str, ...] = ()) -> int | None:
        for i, cell in enumerate(h):
            if _header_cell_matches(cell, *parts, exclude=exclude):
                return i
        return None

    date_i = find("date", "time") or find("date") or 0
    type_i = find("type") or 2
    po_i = (
        find("order", "id")
        or find("purchase", "order")
        or find("po", "number")
        or find("po", "#")
        or 3
    )
    sku_i = find("sku") or 4
    if sku_i == po_i:
        alt = find("sku", exclude=("order", "purchase", "po"))
        sku_i = alt if alt is not None else sku_i
    qty_i = find("quantity") or find("qty") or 6
    amt_i = (
        find("product", "sales")
        or find("total", "sales")
        or find("sales", exclude=("product",))
        or 13
    )
    indices = (date_i, type_i, po_i, sku_i, qty_i, amt_i)
    if indices != _DEFAULT_RAW_COL_INDICES:
        print(f"[amazon] Column map from headers: {indices}", flush=True)
    return indices


def find_header_row_index(rows: list[tuple]) -> int:
    """First row with non-empty values in at least two columns."""
    for i, row in enumerate(rows):
        padded = _pad_row(row, 29)
        filled = sum(1 for v in padded if v is not None and str(v).strip())
        if filled >= 2:
            return i
    raise ValueError(f"Could not find header row (≥2 columns with data) in {len(rows)} rows.")


def _is_output_artifact(path: Path) -> bool:
    """Skip formatted workbooks if they were copied into Input by mistake."""
    return path.stem.strip().endswith(_OUTPUT_STEM_SUFFIX)


def _is_skipped_filename(name: str) -> bool:
    n = name.strip()
    if not n:
        return True
    if n.startswith(_SKIP_NAME_PREFIXES):
        return True
    if n.lower().endswith(".tmp"):
        return True
    return False


def _processed_state_path() -> Path:
    raw = (os.environ.get("AMAZON_PROCESSED_STATE_FILE") or "").strip()
    if raw:
        return Path(raw).expanduser()
    return resolve_amazon_input_dir() / ".amazon_invoice_processed.json"


def _eligible_inputs(folder: Path) -> list[Path]:
    if not folder.is_dir():
        return []
    out: list[Path] = []
    for p in folder.iterdir():
        if not p.is_file():
            continue
        if _is_skipped_filename(p.name):
            continue
        if p.suffix.lower() not in _ELIGIBLE_SUFFIXES:
            continue
        if _is_output_artifact(p):
            continue
        out.append(p)
    return sorted(out, key=lambda p: p.stat().st_mtime, reverse=True)


def wait_for_file_stable(path: Path, *, settle_s: float = 2.0, timeout_s: float = 120.0) -> bool:
    """Wait until size/mtime stop changing (file finished saving from browser/Excel)."""
    deadline = time.monotonic() + timeout_s
    last_size = -1
    last_mtime = -1.0
    stable_since: float | None = None
    while time.monotonic() < deadline:
        try:
            st = path.stat()
        except OSError:
            time.sleep(0.5)
            continue
        if st.st_size == last_size and st.st_mtime == last_mtime:
            if stable_since is None:
                stable_since = time.monotonic()
            elif time.monotonic() - stable_since >= settle_s:
                return True
        else:
            last_size = st.st_size
            last_mtime = st.st_mtime
            stable_since = None
        time.sleep(0.5)
    return False


def _load_processed_state() -> dict[str, float]:
    state_path = _processed_state_path()
    if not state_path.is_file():
        return {}
    try:
        raw = json.loads(state_path.read_text(encoding="utf-8"))
        return {str(k): float(v) for k, v in raw.items()}
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return {}


def _save_processed_state(state: dict[str, float]) -> None:
    state_path = _processed_state_path()
    try:
        state_path.parent.mkdir(parents=True, exist_ok=True)
        state_path.write_text(json.dumps(state, indent=0), encoding="utf-8")
    except OSError:
        pass


def _mark_processed(source: Path) -> None:
    key = str(source.resolve())
    state = _load_processed_state()
    state[key] = source.stat().st_mtime
    _save_processed_state(state)


def _already_processed(source: Path) -> bool:
    key = str(source.resolve())
    state = _load_processed_state()
    try:
        return state.get(key) == source.stat().st_mtime
    except OSError:
        return False


def mark_existing_input_files_processed(folder: Path | None = None) -> int:
    """
    Mark every eligible raw file already in Input as processed (no format/print).

    Used when the watcher starts so restarts only pick up files saved after startup.
    """
    folder = folder or resolve_amazon_input_dir()
    marked = 0
    for path in _eligible_inputs(folder):
        if _already_processed(path):
            continue
        _mark_processed(path)
        marked += 1
    return marked


def _select_columns(
    header: tuple,
    data_rows: list[tuple],
    raw_indices: tuple[int, ...] | None = None,
) -> tuple[list[str], list[tuple]]:
    indices = raw_indices or _DEFAULT_RAW_COL_INDICES
    width = max(29, max(_row_width(header), max((_row_width(r) for r in data_rows), default=0)))
    header = _pad_row(header, width)
    headers_out: list[str] = []
    for idx in indices:
        if idx >= len(header):
            headers_out.append("")
        else:
            headers_out.append("" if header[idx] is None else str(header[idx]).strip())

    rows_out: list[tuple] = []
    for row in data_rows:
        row = _pad_row(row, width)
        if not any(v is not None and str(v).strip() for v in row):
            continue
        picked = tuple(row[i] if i < len(row) else None for i in indices)
        rows_out.append(picked)
    return headers_out, rows_out


def _type_column_index(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if "order" in h.lower() and "refund" not in h.lower():
            return i
        if h.strip().lower() in ("type", "transaction type"):
            return i
    return 1  # raw column C → output B


def _forward_fill_transaction_dates(rows: list[tuple]) -> list[tuple]:
    """Amazon exports often leave date/time blank on continuation lines under a settlement group."""
    filled: list[tuple] = []
    last_date_cell: object = None
    for row in rows:
        if not row:
            filled.append(row)
            continue
        cell = row[0]
        if cell is not None and str(cell).strip():
            last_date_cell = cell
            filled.append(row)
        elif last_date_cell is not None:
            mutable = list(row)
            mutable[0] = last_date_cell
            filled.append(tuple(mutable))
        else:
            filled.append(row)
    return filled


def _forward_fill_po_and_sku(rows: list[tuple]) -> list[tuple]:
    """
    Continuation Order lines often repeat quantity on new rows with blank PO/SKU.

    Without this, merge keys differ (filled PO+SKU vs blank+blank) and identical orders
    stay split as one row per unit.
    """
    filled: list[tuple] = []
    last_po: str = ""
    last_sku: str = ""
    for row in rows:
        if not row:
            filled.append(row)
            continue
        mutable = list(row)
        while len(mutable) <= _COL_SKU:
            mutable.append(None)
        po = _merge_key_po(mutable[_COL_PO])
        sku = _merge_key_sku(mutable[_COL_SKU])
        if po:
            last_po = po
        elif last_po:
            mutable[_COL_PO] = last_po
        if sku:
            last_sku = sku
        elif last_sku:
            mutable[_COL_SKU] = last_sku
        filled.append(tuple(mutable))
    return filled


def _filter_rows(
    headers: list[str],
    rows: list[tuple],
    *,
    keep_dates: set[date],
) -> list[tuple]:
    type_ix = _type_column_index(headers)
    kept: list[tuple] = []
    for row in rows:
        type_val = "" if type_ix >= len(row) or row[type_ix] is None else str(row[type_ix]).strip()
        if type_val.lower() != "order":
            continue
        if not _row_matches_keep_dates(row, keep_dates):
            continue
        kept.append(row)
    return kept


def _sort_rows_by_transaction_datetime(rows: list[tuple]) -> list[tuple]:
    def sort_key(row: tuple) -> tuple[bool, datetime]:
        dt = parse_amazon_transaction_datetime(row[0] if row else None)
        return (dt is None, dt or datetime.max)

    return sorted(rows, key=sort_key)


# Output columns after _select_columns: A date, B type, C PO, D SKU, E qty, F amount.
_COL_PO = 2
_COL_SKU = 3
_COL_QTY = 4
_COL_AMT = 5


def _cell_key(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_merge_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value) and abs(value) < 9_007_199_254_740_992:
            return str(int(value))
        if abs(value) >= 1e11:
            return format(value, ".0f")
        return str(value).strip()
    s = unicodedata.normalize("NFKC", str(value))
    s = s.replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _merge_key_po(value: object) -> str:
    """
    Canonical PO for grouping — rows can look identical but differ by dashes,
    spaces, Excel numeric formatting, or hidden characters.
    """
    s = _normalize_merge_text(value).replace(",", "")
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    try:
        f = float(s)
        if f == int(f) and abs(f) < 9_007_199_254_740_992:
            s = str(int(f))
    except ValueError:
        pass
    # Amazon order IDs: compare by digits so 114-1234567-1234567 == 11412345671234567
    if re.fullmatch(r"[\d\s\-]+", s):
        digits = re.sub(r"\D", "", s)
        if len(digits) >= 8:
            return digits
    compact = re.sub(r"\s+", "", s).upper()
    return compact


def _merge_key_sku(value: object) -> str:
    s = _normalize_merge_text(value)
    return s.casefold() if s else ""


def _parse_quantity(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_amount_for_sum(value: object) -> float:
    coerced = _coerce_accounting_cell_value(value)
    if isinstance(coerced, (int, float)):
        return float(coerced)
    return 0.0


def _quantity_output_value(total: float) -> int | float:
    if total == int(total):
        return int(total)
    return total


def _consolidate_rows_by_po_and_sku(
    rows: list[tuple],
    *,
    po_col: int = _COL_PO,
    sku_col: int = _COL_SKU,
    qty_col: int = _COL_QTY,
    amt_col: int = _COL_AMT,
) -> list[tuple]:
    """
    Merge Order lines with the same PO + SKU (sums quantity and product sales).

    Uses canonical merge keys so three identical-looking unit rows always become one line.
    """
    if not rows:
        return rows

    groups: dict[tuple[str, str], dict] = {}
    order: list[tuple[str, str]] = []
    max_col = max(po_col, sku_col, qty_col, amt_col)

    for i, row in enumerate(rows):
        padded = tuple(row) + (None,) * max(0, max_col + 1 - len(row))
        po_key = _merge_key_po(padded[po_col])
        sku_key = _merge_key_sku(padded[sku_col])
        if not po_key:
            po_key = _normalize_merge_text(padded[po_col]) or f"__row_{i}"
        if not sku_key:
            sku_key = _normalize_merge_text(padded[sku_col]) or f"__row_{i}"
        key = (po_key, sku_key)
        qty = _parse_quantity(padded[qty_col])
        amt = _parse_amount_for_sum(padded[amt_col])

        if key not in groups:
            groups[key] = {"row": padded, "qty": 0.0, "amt": 0.0}
            order.append(key)
        groups[key]["qty"] += qty
        groups[key]["amt"] += amt

    merged: list[tuple] = []
    for key in order:
        g = groups[key]
        row = list(g["row"])
        while len(row) <= amt_col:
            row.append(None)
        row[qty_col] = _quantity_output_value(g["qty"])
        row[amt_col] = g["amt"]
        merged.append(tuple(row))
    return merged


def _log_unmerged_po_sku_groups(
    rows: list[tuple],
    consolidated: list[tuple],
    *,
    po_col: int = _COL_PO,
    sku_col: int = _COL_SKU,
) -> None:
    """Warn when the same visible PO+SKU still appears on multiple output rows."""
    if len(consolidated) >= len(rows):
        return
    out_keys: dict[tuple[str, str], int] = {}
    for row in consolidated:
        pk = (_merge_key_po(row[po_col] if po_col < len(row) else None),
              _merge_key_sku(row[sku_col] if sku_col < len(row) else None))
        out_keys[pk] = out_keys.get(pk, 0) + 1
    dup = [k for k, n in out_keys.items() if n > 1]
    if dup:
        print(
            f"[amazon] WARN: {len(dup)} PO+SKU group(s) still split across multiple output rows "
            f"after consolidate (check raw file for mismatched SKU text).",
            flush=True,
        )


def _order_rows_by_calendar_date(
    rows: list[tuple],
    headers: list[str],
) -> dict[date, list[datetime]]:
    """Count Order rows per settlement calendar date in the raw export (for logging)."""
    type_ix = _type_column_index(headers)
    by_date: dict[date, list[datetime]] = {}
    for row in rows:
        type_val = "" if type_ix >= len(row) or row[type_ix] is None else str(row[type_ix]).strip()
        if type_val.lower() != "order":
            continue
        tx_date = parse_amazon_transaction_date(row[0] if row else None)
        tx_dt = parse_amazon_transaction_datetime(row[0] if row else None)
        if tx_date is None or tx_dt is None:
            continue
        by_date.setdefault(tx_date, []).append(tx_dt)
    return by_date


def _log_source_date_coverage(
    rows: list[tuple],
    headers: list[str],
    keep_dates: set[date],
    run_day: date,
) -> None:
    """Explain what Amazon actually exported for each keep-day (e.g. last time is not a script cut-off)."""
    by_date = _order_rows_by_calendar_date(rows, headers)
    for d in sorted(keep_dates):
        times = sorted(by_date.get(d, []))
        if not times:
            print(f"[amazon]   Source export: 0 Order rows dated {d.isoformat()}", flush=True)
            continue
        print(
            f"[amazon]   Source export: {len(times)} Order row(s) dated {d.isoformat()} "
            f"(times {times[0].strftime('%I:%M:%S %p')} – {times[-1].strftime('%I:%M:%S %p')}; "
            f"all times on that calendar date are kept)",
            flush=True,
        )

    if run_day.weekday() == 0 or _include_prior_settlement_day():
        return
    yesterday = run_day - timedelta(days=1)
    prior = run_day - timedelta(days=2)
    y_count = len(by_date.get(yesterday, []))
    p_count = len(by_date.get(prior, []))
    if p_count > y_count:
        print(
            f"[amazon] NOTE: This file has {p_count} Order rows dated {prior.isoformat()} but only "
            f"{y_count} dated {yesterday.isoformat()}. Amazon often assigns yesterday's sales to the "
            f"prior settlement date. Set AMAZON_INCLUDE_PRIOR_SETTLEMENT_DAY=true (default) to include "
            f"{prior.isoformat()}, or download a newer transaction report after settlements finish.",
            flush=True,
        )


def _coerce_accounting_cell_value(value: object) -> float | str | None:
    """Parse currency/amount text from CSV so column F can use Excel Accounting format."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    s = s.replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return str(value).strip()


def write_amazon_workbook(headers: list[str], rows: list[tuple], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon"
    _COL_AMOUNT = 6  # column F (raw N)

    for c, name in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=name)

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            if c == _COL_AMOUNT:
                val = _coerce_accounting_cell_value(val)
            ws.cell(row=r, column=c, value=val)

    acct = _accounting_format()
    center = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 31.50
    ws.column_dimensions["B"].width = 10.0
    ws.column_dimensions["C"].width = 25.0
    ws.column_dimensions["D"].width = 18.50
    ws.column_dimensions["E"].width = 8.0
    ws.column_dimensions["F"].width = 13.30

    last_row = max(1, 1 + len(rows))
    for r in range(1, last_row + 1):
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=5).alignment = center
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=_COL_AMOUNT).number_format = acct

    _apply_openpyxl_print_file_and_page_footer(ws)

    return _save_xlsx_or_fallback(wb, out_path)


def process_amazon_export(
    source: Path,
    *,
    run_day: date | None = None,
    output_dir: Path | None = None,
    print_workbook: bool = True,
) -> Path:
    load_project_dotenv()
    run_day = run_day or date.today()
    source = source.resolve()
    keep_dates = transaction_dates_to_keep(run_day)

    rows = _load_rows(source)
    header_i = find_header_row_index(rows)
    header = rows[header_i]
    body = rows[header_i + 1 :]

    raw_indices = _resolve_raw_col_indices(header)
    headers, selected = _select_columns(header, body, raw_indices)
    selected = _forward_fill_transaction_dates(selected)
    selected = _forward_fill_po_and_sku(selected)
    print(f"[amazon] Date filter (calendar days only): {', '.join(sorted(d.isoformat() for d in keep_dates))}", flush=True)
    _log_source_date_coverage(selected, headers, keep_dates, run_day)
    filtered = _filter_rows(headers, selected, keep_dates=keep_dates)
    filtered = _sort_rows_by_transaction_datetime(filtered)
    filtered = _forward_fill_po_and_sku(filtered)
    line_count = len(filtered)
    consolidated = _consolidate_rows_by_po_and_sku(filtered)
    _log_unmerged_po_sku_groups(filtered, consolidated)
    if len(consolidated) < line_count:
        print(
            f"[amazon] Consolidated {line_count} Order line(s) -> {len(consolidated)} "
            f"by PO+SKU (columns {headers[_COL_PO]!r} + {headers[_COL_SKU]!r}); "
            "summed quantity and sales.",
            flush=True,
        )
    elif line_count > 0:
        print(
            f"[amazon] No rows merged ({line_count} Order line(s) — each PO+SKU was already unique).",
            flush=True,
        )

    out_dir = (output_dir or resolve_amazon_output_dir()).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_path_for_source(source, out_dir)

    saved = write_amazon_workbook(headers, consolidated, out_path)
    date_list = ", ".join(sorted(d.isoformat() for d in keep_dates))
    print(
        f"[amazon] {source.name}: {len(selected)} data row(s) in export -> "
        f"kept {line_count} Order row(s) for {date_list} -> {len(consolidated)} output row(s) -> {saved}",
        flush=True,
    )

    if not consolidated:
        print(
            f"[amazon] WARNING: No rows matched the date filter ({date_list}). "
            "Report saved with headers only; print skipped. "
            "Check transaction dates in column A or run on the correct weekday.",
            flush=True,
        )
        print_workbook = False

    if print_workbook:
        try:
            from depot_excel_print import print_landscape_with_gridlines

            print_landscape_with_gridlines(saved)
            print(f"[amazon] Print sent for {saved.name}", flush=True)
        except Exception as e:
            import traceback

            print(f"[amazon] Excel print failed (workbook saved): {e}", flush=True)
            traceback.print_exc()

    _mark_processed(source)
    return saved


def _skip_reason(path: Path) -> str:
    if _is_skipped_filename(path.name):
        return "ignored (temp/hidden filename)"
    if path.suffix.lower() not in _ELIGIBLE_SUFFIXES:
        return f"ignored (extension {path.suffix!r} - need .xlsx, .xlsm, or .csv)"
    if _is_output_artifact(path):
        return "ignored (formatted Output workbook — belongs in Output folder)"
    if _already_processed(path):
        return (
            "already processed (delete "
            f"{_processed_state_path()} or save under a new name)"
        )
    return "eligible"


def describe_folder_state(folder: Path | None = None) -> str:
    """Human-readable summary for logs when nothing runs."""
    folder = folder or resolve_amazon_input_dir()
    if not folder.is_dir():
        return f"folder missing: {folder}"
    eligible = _eligible_inputs(folder)
    if not eligible:
        names = [p.name for p in folder.iterdir() if p.is_file()][:12]
        return f"no raw .xlsx/.csv in {folder}" + (f" (files seen: {names})" if names else " (empty)")
    newest = eligible[0]
    reason = _skip_reason(newest)
    if reason != "eligible":
        return f"newest file {newest.name}: {reason}"
    return f"ready: {newest.name}"


def folder_file_snapshot(folder: Path) -> dict[str, float]:
    """All files in folder (any extension) -> mtime, for change detection."""
    snap: dict[str, float] = {}
    if not folder.is_dir():
        return snap
    for p in folder.iterdir():
        if not p.is_file():
            continue
        try:
            snap[p.name] = p.stat().st_mtime
        except OSError:
            continue
    return snap


def log_folder_scan(folder: Path) -> None:
    """Print one line per file so logs show what the watcher actually sees."""
    folder = folder.resolve()
    print(f"[amazon] Folder scan: {folder}", flush=True)
    if not folder.is_dir():
        print("[amazon]   (folder not accessible)", flush=True)
        return
    files = sorted(folder.iterdir(), key=lambda p: p.name.lower())
    file_items = [p for p in files if p.is_file()]
    if not file_items:
        print("[amazon]   (no files in folder)", flush=True)
        return
    for p in file_items:
        print(f"[amazon]   {p.name}: {_skip_reason(p)}", flush=True)


def pick_newest_unprocessed(folder: Path | None = None) -> Path | None:
    folder = folder or resolve_amazon_input_dir()
    for p in _eligible_inputs(folder):
        if not _already_processed(p):
            return p
    return None


def process_newest_in_folder(folder: Path | None = None, *, force: bool = False) -> Path | None:
    folder = folder or resolve_amazon_input_dir()
    if force:
        candidates = _eligible_inputs(folder)
        if not candidates:
            return None
        return process_amazon_export(candidates[0])
    path = pick_newest_unprocessed(folder)
    if path is None:
        return None
    return process_amazon_export(path)


def watch_folder(
    folder: Path | None = None,
    *,
    interval_s: float = 30.0,
) -> None:
    folder = folder or resolve_amazon_input_dir()
    print(f"[amazon] Watching {folder} every {interval_s:.0f}s for new exports…", flush=True)
    while True:
        try:
            path = pick_newest_unprocessed(folder)
            if path is not None:
                process_amazon_export(path)
        except Exception as e:
            print(f"[amazon] ERROR: {e}", flush=True)
        time.sleep(interval_s)


def main(argv: list[str] | None = None) -> int:
    load_project_dotenv()
    parser = argparse.ArgumentParser(description="Format and print Amazon invoice exports from the share.")
    parser.add_argument(
        "path",
        nargs="?",
        help="Raw Amazon export (.xlsx/.csv). If omitted, process newest unprocessed file in the Amazon folder.",
    )
    parser.add_argument("--watch", action="store_true", help="Poll the Amazon folder for new files.")
    parser.add_argument("--force", action="store_true", help="Re-process newest raw file even if already done.")
    parser.add_argument("--interval", type=float, default=30.0, help="Watch poll interval in seconds.")
    parser.add_argument("--no-print", action="store_true", help="Save only; skip Excel print.")
    args = parser.parse_args(argv)

    if (os.environ.get("AMAZON_INVOICE_POSTPROCESS") or "true").strip().lower() in ("0", "false", "no"):
        print("[amazon] Skipped (AMAZON_INVOICE_POSTPROCESS is false).", flush=True)
        return 0

    if args.watch:
        watch_folder(interval_s=args.interval)
        return 0

    if args.path:
        process_amazon_export(
            Path(args.path),
            print_workbook=not args.no_print,
        )
        return 0

    result = process_newest_in_folder(force=args.force)
    if result is None:
        folder = resolve_amazon_input_dir()
        print(f"[amazon] No new raw export to process in {folder}", flush=True)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
