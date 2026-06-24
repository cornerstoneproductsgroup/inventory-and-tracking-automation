"""Read CornerstoneMaster order rows for WorldShip label naming."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from automation.worldship_label_config import (
    COL_LABEL_PR,
    COL_PO,
    COL_RETAILER,
    COL_SKU,
    CORNERSTONE_MASTER_DIR,
    CORNERSTONE_MASTER_DIR_FALLBACK,
    DATA_START_ROW,
    retailer_merchant_to_key,
)

SKU_HEADER_HINTS = ("sku",)
PO_HEADER_HINTS = (
    "purchase_order_number",
    "purchase_order",
    "purchaseordernumber",
    "po_number",
    "ponumber",
    "customer_po",
)
RETAILER_HEADER_HINTS = (
    "merchant_id",
    "merchantid",
    "merchant",
    "retailer",
)
LABEL_PR_HEADER_HINTS = (
    "label_pr",
    "label_profile",
    "labelprinter",
    "label_printer",
)


def _log(msg: str) -> None:
    print(f"[worldship] {msg}", flush=True)


@dataclass(frozen=True)
class CornerstoneOrderRow:
    row_number: int
    sku: str
    po: str
    retailer_key: str
    retailer_raw: str
    label_pr: str = ""


def normalize_label_pr(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").strip()).lower()


def is_cornerstone_warehouse_print_row(label_pr: str) -> bool | None:
    """
    Return True/False from LABEL_PR (LabelPDF vs Label1), or None when blank
    so callers can fall back to the warehouse vendor list.
    """
    norm = normalize_label_pr(label_pr)
    if not norm:
        return None
    if norm == "labelpdf" or norm.endswith("pdf"):
        return False
    if norm == "label1" or norm.startswith("label1"):
        return True
    if "print" in norm and "pdf" not in norm:
        return True
    return False


def _resolve_master_dir() -> Path:
    for cand in (CORNERSTONE_MASTER_DIR, CORNERSTONE_MASTER_DIR_FALLBACK):
        try:
            if cand.is_dir():
                return cand
        except OSError:
            continue
    raise FileNotFoundError(
        f"CornerstoneMaster folder not found. Checked:\n"
        f"  {CORNERSTONE_MASTER_DIR}\n"
        f"  {CORNERSTONE_MASTER_DIR_FALLBACK}\n"
        "Set WORLDSHIP_CORNERSTONE_MASTER_DIR in Inventory Submissions\\.env"
    )


def _find_master_file(folder: Path) -> Path:
    patterns = (
        "CornerstoneMaster.csv",
        "CornerstoneMaster*.csv",
        "CornerstoneMaster*.xlsx",
        "CornerstoneMaster*.xlsm",
        "Cornerstone Master*.csv",
        "Cornerstone Master*.xlsx",
    )
    matches: list[Path] = []
    for pattern in patterns:
        matches.extend(folder.glob(pattern))
    if not matches:
        raise FileNotFoundError(
            f"No CornerstoneMaster file found in {folder}.\n"
            "  Expected CornerstoneMaster.csv or CornerstoneMaster.xlsx"
        )
    exact_csv = folder / "CornerstoneMaster.csv"
    if exact_csv in matches:
        chosen = exact_csv.resolve()
    else:
        matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        chosen = matches[0].resolve()
    _log(f"Using CornerstoneMaster file: {chosen}")
    return chosen


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _purchase_order_for_label(raw: str) -> str:
    """
    Full PURCHASE_ORDER cell for the saved PDF name (e.g. '48690515 Coarse 10').
    Invalid Windows filename characters are replaced with underscores.
    """
    text = _cell_str(raw)
    if not text:
        return ""
    text = re.sub(r'[<>:"/\\|?*]', "_", text)
    return re.sub(r"\s+", " ", text).strip()


def _norm_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (name or "").strip().lower()).strip("_")


def _header_index(header: list[str], hints: tuple[str, ...]) -> int | None:
    """Match column by exact header name first; avoid short hints like 'po' in SHPTO_*."""
    normalized = [(i, _norm_header(raw)) for i, raw in enumerate(header)]
    for hint in hints:
        h = _norm_header(hint)
        for i, norm in normalized:
            if norm == h:
                return i
    for hint in hints:
        h = _norm_header(hint)
        if len(h) < 5:
            continue
        for i, norm in normalized:
            if h in norm or norm in h:
                return i
    return None


def _label_pr_column_index(header: list[str]) -> int:
    label_i = _header_index(header, LABEL_PR_HEADER_HINTS)
    if label_i is not None:
        return label_i
    return column_index_from_string(COL_LABEL_PR) - 1


def _looks_like_label_pr_value(value: str) -> bool:
    norm = normalize_label_pr(value)
    if not norm:
        return False
    if norm in ("labelpdf", "label1"):
        return True
    if norm.endswith("pdf") or norm.startswith("label1"):
        return True
    return "print" in norm and "pdf" not in norm


def _label_pr_column_score(data_rows: list[list[str]], col_i: int) -> int:
    score = 0
    for row in data_rows[:80]:
        if col_i < 0:
            continue
        cell = _cell_str(row[col_i] if len(row) > col_i else "")
        if _looks_like_label_pr_value(cell):
            score += 1
    return score


def _resolve_label_pr_column_index(
    header: list[str],
    data_rows: list[list[str]],
    *,
    default_i: int,
) -> int:
    """
    Pick the column that actually contains LabelPDF / Label1 values.

    WorldShip CSVs often have an empty column at the env default while LABEL_PR
    data lives in another column (commonly X). When the default column is blank,
    scanning row content avoids misclassifying rows via the warehouse-vendor fallback.
    """
    sample = data_rows[:80]
    ncol = len(header)
    for row in sample:
        ncol = max(ncol, len(row))

    header_i = _header_index(header, LABEL_PR_HEADER_HINTS)
    env_i = column_index_from_string(COL_LABEL_PR) - 1

    best_i = default_i
    best_score = _label_pr_column_score(sample, default_i)

    for col_i in range(ncol):
        score = _label_pr_column_score(sample, col_i)
        if score > best_score:
            best_score = score
            best_i = col_i

    if best_score > 0 and best_i != default_i:
        hdr = header[best_i] if best_i < len(header) else ""
        _log(
            f"LABEL_PR column auto-detected: col {best_i + 1} "
            f"({hdr!r} or {COL_LABEL_PR} default col {default_i + 1} was blank). "
            f"Found {best_score} LabelPDF/Label1 value(s) in sample rows."
        )
    elif best_score == 0 and header_i is None:
        hdr = header[default_i] if default_i < len(header) else ""
        if not hdr:
            _log(
                f"WARN: No LabelPDF/Label1 values found in CornerstoneMaster sample; "
                f"using col {default_i + 1} ({COL_LABEL_PR}) for LABEL_PR. "
                f"Blank LABEL_PR falls back to warehouse-vendor list for print vs save."
            )

    return best_i


def _column_indices(header: list[str]) -> tuple[int, int, int, int]:
    """SKU, PO, retailer, LABEL_PR column indices (0-based). Prefer named CSV headers."""
    sku_i = _header_index(header, SKU_HEADER_HINTS)
    po_i = _header_index(header, PO_HEADER_HINTS)
    retailer_i = _header_index(header, RETAILER_HEADER_HINTS)
    label_i = _label_pr_column_index(header)
    if sku_i is not None and po_i is not None and retailer_i is not None:
        return sku_i, po_i, retailer_i, label_i
    _log(
        "WARN: CornerstoneMaster header names not recognized — "
        f"using Excel columns {COL_SKU}/{COL_PO}/{COL_RETAILER}/{COL_LABEL_PR}."
    )
    return (
        column_index_from_string(COL_SKU) - 1,
        column_index_from_string(COL_PO) - 1,
        column_index_from_string(COL_RETAILER) - 1,
        column_index_from_string(COL_LABEL_PR) - 1,
    )


def _find_header_row_index(all_rows: list[list[str]]) -> int:
    """Locate the header row (WorldShip CSV may use row 1)."""
    for idx, row in enumerate(all_rows[:15]):
        header = [str(c).strip() for c in row]
        if not any(header):
            continue
        if _header_index(header, SKU_HEADER_HINTS) is None:
            continue
        if _header_index(header, PO_HEADER_HINTS) is None:
            continue
        if _header_index(header, RETAILER_HEADER_HINTS) is None:
            continue
        return idx
    return 0


def _read_csv_rows(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "latin1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f"Could not decode CSV: {path}")

    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        if text.count("\t") > text.count(","):
            delimiter = "\t"

    return list(csv.reader(text.splitlines(), delimiter=delimiter))


def _append_order_row(
    rows: list[CornerstoneOrderRow],
    *,
    row_idx: int,
    row: tuple | list,
    sku_i: int,
    po_i: int,
    retailer_i: int,
    label_pr_i: int,
) -> str:
    """
    Append one order row.

    Returns 'added', 'skip' (blank line), or 'stop' (end of data).
    """
    if not row or not any(_cell_str(c) for c in row):
        return "stop"
    sku = _cell_str(row[sku_i] if len(row) > sku_i else "")
    if not sku:
        return "skip"
    po = _purchase_order_for_label(row[po_i] if len(row) > po_i else "")
    retailer_raw = _cell_str(row[retailer_i] if len(row) > retailer_i else "")
    label_pr = _cell_str(row[label_pr_i] if len(row) > label_pr_i else "")
    if not po:
        raise ValueError(f"Row {row_idx}: PO is empty (SKU={sku!r}).")
    if not retailer_raw:
        raise ValueError(f"Row {row_idx}: retailer/merchant is empty.")
    rows.append(
        CornerstoneOrderRow(
            row_number=row_idx,
            sku=sku,
            po=po,
            retailer_key=retailer_merchant_to_key(retailer_raw),
            retailer_raw=retailer_raw,
            label_pr=label_pr,
        )
    )
    return "added"


def _load_csv(path: Path, *, limit: int | None) -> list[CornerstoneOrderRow]:
    all_rows = _read_csv_rows(path)
    if not all_rows:
        raise ValueError(f"CSV is empty: {path}")

    header_idx = _find_header_row_index(all_rows)
    header = [str(c).strip() for c in all_rows[header_idx]]
    sku_i, po_i, retailer_i, default_label_i = _column_indices(header)
    data_start = max(header_idx + 1, DATA_START_ROW - 1)
    data_rows = all_rows[data_start:]
    label_pr_i = _resolve_label_pr_column_index(
        header, data_rows, default_i=default_label_i
    )
    label_hdr = header[label_pr_i] if label_pr_i < len(header) else COL_LABEL_PR
    _log(
        "CornerstoneMaster columns: "
        f"SKU={header[sku_i]!r} (col {sku_i + 1}), "
        f"PO={header[po_i]!r} (col {po_i + 1}), "
        f"retailer={header[retailer_i]!r} (col {retailer_i + 1}), "
        f"LABEL_PR={label_hdr!r} (col {label_pr_i + 1})"
    )

    rows: list[CornerstoneOrderRow] = []
    for offset, row in enumerate(data_rows, start=data_start + 1):
        if limit is not None and len(rows) >= limit:
            break
        result = _append_order_row(
            rows,
            row_idx=offset,
            row=row,
            sku_i=sku_i,
            po_i=po_i,
            retailer_i=retailer_i,
            label_pr_i=label_pr_i,
        )
        if result == "stop":
            break
    return rows


def _load_xlsx(path: Path, *, limit: int | None) -> list[CornerstoneOrderRow]:
    rows: list[CornerstoneOrderRow] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(c or "").strip() for c in (first_row or ())]
        if _header_index(header, SKU_HEADER_HINTS) is not None:
            sku_i, po_i, retailer_i, default_label_i = _column_indices(header)
            start_row = 2
        else:
            sku_i, po_i, retailer_i, default_label_i = _column_indices([])
            start_row = DATA_START_ROW

        raw_data: list[tuple[int, tuple]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True),
            start=start_row,
        ):
            if not row or not any(_cell_str(c) for c in row):
                break
            raw_data.append((row_idx, row))

        label_pr_i = _resolve_label_pr_column_index(
            header,
            [list(r) for _, r in raw_data],
            default_i=default_label_i,
        )
        label_hdr = header[label_pr_i] if label_pr_i < len(header) else COL_LABEL_PR
        _log(
            "CornerstoneMaster columns: "
            f"SKU col {sku_i + 1}, PO col {po_i + 1}, retailer col {retailer_i + 1}, "
            f"LABEL_PR={label_hdr!r} (col {label_pr_i + 1})"
        )

        for row_idx, row in raw_data:
            if limit is not None and len(rows) >= limit:
                break
            result = _append_order_row(
                rows,
                row_idx=row_idx,
                row=row,
                sku_i=sku_i,
                po_i=po_i,
                retailer_i=retailer_i,
                label_pr_i=label_pr_i,
            )
            if result == "stop":
                break
    finally:
        wb.close()
    return rows


def load_cornerstone_orders(
    *,
    limit: int | None = None,
    master_path: Path | None = None,
) -> list[CornerstoneOrderRow]:
    path = master_path or _find_master_file(_resolve_master_dir())
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _load_csv(path, limit=limit)
    elif suffix in (".xlsx", ".xlsm"):
        rows = _load_xlsx(path, limit=limit)
    else:
        raise ValueError(f"Unsupported CornerstoneMaster format: {path}")

    if not rows:
        raise ValueError(
            f"No order rows found in {path}. "
            "Expected headers SKU, PURCHASE_ORDER_NUMBER, and MERCHANT_ID with data below."
        )
    _log(f"Loaded {len(rows)} CornerstoneMaster row(s) from {path.name}.")
    return rows
