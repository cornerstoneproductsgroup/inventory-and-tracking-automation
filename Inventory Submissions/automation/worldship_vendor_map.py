"""SKU → vendor folder mapping for WorldShip label saves."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from automation.worldship_label_config import (
    DEFAULT_VENDOR_MAP_CANDIDATES,
    RETAILER_VENDOR_MAP_FILES,
    VENDOR_MAP_DIR,
)


def _log(msg: str) -> None:
    print(f"[worldship] {msg}", flush=True)


def resolve_vendor_map_path_for_retailer(retailer_key: str) -> Path:
    """Return the Order Splitter vendor map for a retailer (depot, thdso, tractor, …)."""
    filename = RETAILER_VENDOR_MAP_FILES.get(retailer_key)
    if not filename:
        raise ValueError(
            f"No vendor map configured for retailer key {retailer_key!r}. "
            f"Known keys: {', '.join(sorted(RETAILER_VENDOR_MAP_FILES))}"
        )
    path = VENDOR_MAP_DIR / filename
    try:
        if path.is_file():
            return path.resolve()
    except OSError:
        pass
    raise FileNotFoundError(
        f"Vendor map not found: {path}\n"
        f"  Expected {filename!r} for retailer {retailer_key!r} in:\n"
        f"  {VENDOR_MAP_DIR}\n"
        f"  Set WORLDSHIP_VENDOR_MAP_DIR if the maps folder is elsewhere."
    )


def resolve_vendor_map_path() -> Path:
    """Legacy single-file lookup (WORLDSHIP_VENDOR_MAP or share fallbacks)."""
    for cand in DEFAULT_VENDOR_MAP_CANDIDATES:
        if not cand:
            continue
        try:
            if cand.is_file():
                return cand.resolve()
        except OSError:
            continue
    tried = [str(c) for c in DEFAULT_VENDOR_MAP_CANDIDATES if c]
    raise FileNotFoundError(
        "Could not find a SKU → vendor mapping file.\n"
        f"  Expected vendor_map_*.xlsx in {VENDOR_MAP_DIR}\n"
        "  or set WORLDSHIP_VENDOR_MAP to a single map file.\n"
        f"  Tried: {', '.join(tried) or '(none)'}"
    )


def _norm_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (name or "").strip().lower()).strip("_")


SKU_COLUMN_HINTS = (
    "sku",
    "model_number",
    "model",
    "model_no",
    "model_num",
    "item",
    "part",
    "style",
    "product",
)
VENDOR_COLUMN_HINTS = (
    "vendor",
    "vendor_name",
    "brand",
    "company",
    "folder",
)


def _column_index(header: list[str], hints: tuple[str, ...]) -> int | None:
    normalized = [(i, _norm_header(raw)) for i, raw in enumerate(header)]
    for hint in hints:
        h = _norm_header(hint)
        for i, norm in normalized:
            if norm == h:
                return i
    for hint in hints:
        h = _norm_header(hint)
        if len(h) < 4:
            continue
        for i, norm in normalized:
            if h in norm or norm in h:
                return i
    return None


def _pick_columns(header: list[str]) -> tuple[int, int]:
    sku_idx = _column_index(header, SKU_COLUMN_HINTS)
    vendor_idx = _column_index(header, VENDOR_COLUMN_HINTS)
    if sku_idx is None or vendor_idx is None:
        raise ValueError(
            f"Vendor map needs Model Number (SKU) and Vendor columns; got: {header!r}"
        )
    _log(
        f"Vendor map columns: SKU={header[sku_idx]!r}, vendor={header[vendor_idx]!r}"
    )
    return sku_idx, vendor_idx


def _load_csv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return out
    sku_idx, vendor_idx = _pick_columns(rows[0])
    for row in rows[1:]:
        if len(row) <= max(sku_idx, vendor_idx):
            continue
        sku = (row[sku_idx] or "").strip().upper()
        vendor = (row[vendor_idx] or "").strip()
        if sku and vendor:
            out[sku] = vendor
    return out


def _load_xlsx(path: Path) -> dict[str, str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return {}
    header = [str(c or "").strip() for c in rows[0]]
    sku_idx, vendor_idx = _pick_columns(header)
    out: dict[str, str] = {}
    for row in rows[1:]:
        if not row or len(row) <= max(sku_idx, vendor_idx):
            continue
        sku = str(row[sku_idx] or "").strip().upper()
        vendor = str(row[vendor_idx] or "").strip()
        if sku and vendor:
            out[sku] = vendor
    return out


def load_vendor_map(path: Path | None = None, *, retailer_key: str | None = None) -> dict[str, str]:
    if path is None and retailer_key:
        src = resolve_vendor_map_path_for_retailer(retailer_key)
    else:
        src = path or resolve_vendor_map_path()
    suffix = src.suffix.lower()
    if suffix == ".csv":
        mapping = _load_csv(src)
    elif suffix in (".xlsx", ".xlsm"):
        mapping = _load_xlsx(src)
    else:
        raise ValueError(f"Unsupported vendor map format: {src}")
    if not mapping:
        raise ValueError(f"Vendor map is empty: {src}")
    _log(f"Loaded {len(mapping)} SKU → vendor entries from {src.name}")
    return mapping


def _compact_sku_key(sku: str) -> str:
    """Ignore spaces when matching (FedEx output often drops spaces in model numbers)."""
    return re.sub(r"\s+", "", (sku or "").strip().upper())


def is_sku_in_vendor_map(sku: str, mapping: dict[str, str]) -> bool:
    """True when SKU matches the vendor map (exact or longest-prefix match)."""
    key = (sku or "").strip().upper()
    if not key:
        return False
    if key in mapping:
        return True
    for prefix in sorted(mapping, key=len, reverse=True):
        if key.startswith(prefix):
            return True
    compact = _compact_sku_key(key)
    if not compact:
        return False
    for map_sku in mapping:
        if _compact_sku_key(map_sku) == compact:
            return True
        if compact.startswith(_compact_sku_key(map_sku)):
            return True
    return False


def lookup_vendor_folder(sku: str, mapping: dict[str, str]) -> str:
    key = (sku or "").strip().upper()
    if not key:
        raise ValueError("SKU is empty.")
    if key in mapping:
        return mapping[key]
    for prefix in sorted(mapping, key=len, reverse=True):
        if key.startswith(prefix):
            return mapping[prefix]

    compact = _compact_sku_key(key)
    if compact:
        for map_sku, vendor in mapping.items():
            if _compact_sku_key(map_sku) == compact:
                if map_sku != key:
                    _log(
                        f"SKU {sku!r} matched vendor map entry {map_sku!r} "
                        f"(ignoring spaces) → {vendor!r}"
                    )
                return vendor
        for map_sku in sorted(mapping, key=lambda k: len(_compact_sku_key(k)), reverse=True):
            map_compact = _compact_sku_key(map_sku)
            if map_compact and compact.startswith(map_compact):
                if map_sku != key:
                    _log(
                        f"SKU {sku!r} prefix-matched map entry {map_sku!r} "
                        f"(ignoring spaces) → {mapping[map_sku]!r}"
                    )
                return mapping[map_sku]

    raise KeyError(f"No vendor mapping for SKU {sku!r}")


def _sku_lookup_keys(sku: str) -> list[str]:
    """Ordered SKU strings to try against the vendor map (exact, + variants, tokens)."""
    raw = (sku or "").strip()
    if not raw:
        return []
    keys: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        key = (value or "").strip().upper()
        if key and key not in seen:
            seen.add(key)
            keys.append(key)

    add(raw)
    upper = raw.upper()
    add(re.sub(r"\s+", "", upper))
    if "+" in upper:
        add(upper.replace("+", ""))
        add(upper.replace("+", " "))
        add(upper.split("+", 1)[0])
        for part in upper.split("+"):
            add(part)
    for token in upper.split():
        add(token)
    return keys


def lookup_vendor_folder_resilient(sku: str, mapping: dict[str, str]) -> str:
    """Try exact/prefix match with + and multi-token SKU variants."""
    last_error: KeyError | None = None
    for key in _sku_lookup_keys(sku):
        try:
            return lookup_vendor_folder(key, mapping)
        except KeyError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("SKU is empty.")


class VendorMapRegistry:
    """Lazy-load Order Splitter vendor maps per retailer."""

    def __init__(self) -> None:
        self._cache: dict[str, dict[str, str]] = {}

    def lookup(self, sku: str, retailer_key: str) -> str:
        if retailer_key not in self._cache:
            self._cache[retailer_key] = load_vendor_map(retailer_key=retailer_key)
        return lookup_vendor_folder(sku, self._cache[retailer_key])
